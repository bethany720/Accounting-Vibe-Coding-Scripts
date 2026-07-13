"""
Prepare_Quarter_Workpapers_V16.py

Based on:
    V15 hybrid approach.

V16 fix:
    - V15 formatting only worked on the first target workbook.
    - V16 makes the Excel formatting pass more explicit and repeatable:
        - Opens PPPROJECTS source workbook once.
        - For each target workbook:
            - Activates the source workbook/sheet before every source copy.
            - Activates the target workbook/sheet before every paste.
            - Re-copies the PPPROJECTS range fresh for each target workbook.
            - Clears Excel clipboard after each workbook.
    - Keeps openpyxl for values/formulas and rename flow.
    - Keeps Excel COM for Paste Special > Formats and worksheet unprotect.

Requirements:
    pip install openpyxl pywin32

Important:
    Close all target workbooks and the ppprojects workbook before running.
"""

from copy import copy
from pathlib import Path
import sys

from openpyxl import load_workbook

try:
    import win32com.client as win32
except ImportError:
    print("ERROR: pywin32 is not installed.")
    print("Install it with: pip install pywin32")
    input("Press Enter to exit...")
    sys.exit(1)


# ----------------------------
# Config
# ----------------------------

EXCLUSIONS_FILE_NAME = "Exclusions.xlsx"
LOOKUP_SHEET_NAME = "Worksheets"

PPPROJECTS_PREFIX = "ppprojects"

SOURCE_SHEET_NAME = "Workpaper"
TARGET_SHEET_NAME = "Department"

SOURCE_COL_G = 7
SOURCE_START_COL_J = 10

ROW_TO_IDENTIFY_END_COLUMN = 6
EXTRA_COLUMNS_TO_COPY = 100
EXTRA_ROWS_TO_COPY = 50

UNPROTECT_PASSWORD = "PPLabor"

EXCEL_EXTENSIONS = [".xlsx", ".xlsm"]

# Excel constants
XL_PASTE_FORMATS = -4122
XL_PASTE_COLUMN_WIDTHS = 8


# ----------------------------
# Basic helpers
# ----------------------------

def normalize(value):
    if value is None:
        return ""
    return str(value).strip().casefold()


def get_yes_no(prompt):
    return input(prompt).strip().casefold() == "yes"


def safe_filename_part(value):
    value = str(value).strip()
    invalid_chars = r'<>:"/\\|?*'
    for ch in invalid_chars:
        value = value.replace(ch, "")
    return value.strip()


def col_letter_openpyxl(ws, col_num):
    return ws.cell(row=1, column=col_num).column_letter


def excel_col_letter(col_num):
    result = ""
    while col_num:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def find_sheet_case_insensitive(workbook, desired_sheet_name):
    desired = normalize(desired_sheet_name)
    for ws in workbook.worksheets:
        if normalize(ws.title) == desired:
            return ws
    return None


def find_excel_sheet_case_insensitive(workbook, desired_sheet_name):
    desired = normalize(desired_sheet_name)
    for ws in workbook.Worksheets:
        if normalize(ws.Name) == desired:
            return ws
    return None


def get_excel_files(folder):
    return [
        p for p in folder.iterdir()
        if p.is_file()
        and p.suffix.casefold() in EXCEL_EXTENSIONS
        and not p.name.startswith("~$")
    ]


# ----------------------------
# Lookup and rename helpers
# ----------------------------

def read_worksheet_lookup_values(exclusions_path):
    wb = load_workbook(exclusions_path, data_only=True)
    ws = find_sheet_case_insensitive(wb, LOOKUP_SHEET_NAME)

    if ws is None:
        wb.close()
        raise ValueError(f"Could not find sheet '{LOOKUP_SHEET_NAME}' in {exclusions_path.name}.")

    values = []
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip():
            values.append(str(val).strip())

    wb.close()

    if not values:
        raise ValueError(f"No lookup values found in Column A of '{LOOKUP_SHEET_NAME}'.")

    values.sort(key=len, reverse=True)
    return values


def find_ppprojects_file(excel_files):
    matches = [
        f for f in excel_files
        if normalize(f.name).startswith(normalize(PPPROJECTS_PREFIX))
    ]

    if len(matches) != 1:
        raise ValueError(
            f"Expected exactly 1 file starting with '{PPPROJECTS_PREFIX}'. Found {len(matches)}."
        )

    return matches[0]


def build_rename_plan(excel_files, ppprojects_file, lookup_values, reporting_period):
    rename_plan = []

    for file_path in excel_files:
        if file_path.resolve() == ppprojects_file.resolve():
            continue

        stem = file_path.stem
        stem_norm = normalize(stem)

        matched_value = None
        matched_index = None

        for lookup_value in lookup_values:
            idx = stem_norm.find(normalize(lookup_value))
            if idx != -1:
                matched_value = lookup_value
                matched_index = idx
                break

        if matched_value is None:
            print(f"WARNING: No lookup match found for file: {file_path.name}")
            continue

        prefix_end = matched_index + len(matched_value)
        new_stem = stem[:prefix_end] + "_" + reporting_period
        new_path = file_path.with_name(new_stem + file_path.suffix)

        if file_path.name == new_path.name:
            rename_plan.append((file_path, new_path, matched_value, "already_named"))
        else:
            rename_plan.append((file_path, new_path, matched_value, "rename"))

    return rename_plan


# ----------------------------
# Source range calculation
# ----------------------------

def existing_cell_has_anything(cell):
    if cell is None:
        return False
    if cell.value is not None:
        return True
    if cell.has_style:
        return True
    if cell.comment is not None:
        return True
    if cell.hyperlink is not None:
        return True
    return False


def get_last_used_or_formatted_col_in_row_no_mutation(ws, row_number):
    last_col = SOURCE_START_COL_J

    for (row, col), cell in ws._cells.items():
        if row == row_number and existing_cell_has_anything(cell):
            if col > last_col:
                last_col = col

    return last_col


def calculate_source_copy_settings(source_ws):
    row_6_last_col = get_last_used_or_formatted_col_in_row_no_mutation(
        source_ws,
        ROW_TO_IDENTIFY_END_COLUMN
    )

    copy_end_col = min(row_6_last_col + EXTRA_COLUMNS_TO_COPY, 16384)
    copy_end_row = source_ws.max_row + EXTRA_ROWS_TO_COPY

    return {
        "row_6_last_col": row_6_last_col,
        "copy_end_col": copy_end_col,
        "copy_end_row": copy_end_row,
    }


# ----------------------------
# openpyxl data copy helpers
# ----------------------------

def copy_value_formula_comment_hyperlink(source_cell, target_cell):
    target_cell.value = source_cell.value

    if source_cell.comment:
        target_cell.comment = copy(source_cell.comment)
    else:
        target_cell.comment = None

    if source_cell.hyperlink:
        target_cell._hyperlink = copy(source_cell.hyperlink)
    else:
        target_cell._hyperlink = None


def copy_column_dimensions_openpyxl(source_ws, target_ws, start_col, end_col):
    for col_idx in range(start_col, end_col + 1):
        col_letter_value = col_letter_openpyxl(source_ws, col_idx)

        source_dim = source_ws.column_dimensions[col_letter_value]
        target_dim = target_ws.column_dimensions[col_letter_value]

        target_dim.width = source_dim.width
        target_dim.hidden = source_dim.hidden
        target_dim.outlineLevel = source_dim.outlineLevel
        target_dim.collapsed = source_dim.collapsed


def copy_row_dimensions_openpyxl(source_ws, target_ws, start_row, end_row):
    for row_idx in range(start_row, end_row + 1):
        source_dim = source_ws.row_dimensions[row_idx]
        target_dim = target_ws.row_dimensions[row_idx]

        target_dim.height = source_dim.height
        target_dim.hidden = source_dim.hidden
        target_dim.outlineLevel = source_dim.outlineLevel
        target_dim.collapsed = source_dim.collapsed


def update_target_values_with_openpyxl(target_file, source_ws, source_copy_settings):
    print(f"\nUpdating values/formulas with openpyxl: {target_file.name}")

    target_wb = load_workbook(target_file)
    target_ws = find_sheet_case_insensitive(target_wb, TARGET_SHEET_NAME)

    if target_ws is None:
        target_wb.close()
        raise ValueError(
            f"{target_file.name} does not contain a sheet named '{TARGET_SHEET_NAME}'."
        )

    copy_end_col = source_copy_settings["copy_end_col"]
    copy_end_row = source_copy_settings["copy_end_row"]

    print(f"  Data range: Column G and Columns J:{col_letter_openpyxl(source_ws, copy_end_col)}")
    print(f"  Data rows: 1:{copy_end_row}")

    # Column G values/formulas.
    for row in range(1, copy_end_row + 1):
        copy_value_formula_comment_hyperlink(
            source_ws.cell(row=row, column=SOURCE_COL_G),
            target_ws.cell(row=row, column=SOURCE_COL_G)
        )

    # Columns J:end values/formulas.
    for row in range(1, copy_end_row + 1):
        for col in range(SOURCE_START_COL_J, copy_end_col + 1):
            copy_value_formula_comment_hyperlink(
                source_ws.cell(row=row, column=col),
                target_ws.cell(row=row, column=col)
            )

    copy_column_dimensions_openpyxl(source_ws, target_ws, SOURCE_COL_G, SOURCE_COL_G)
    copy_column_dimensions_openpyxl(source_ws, target_ws, SOURCE_START_COL_J, copy_end_col)
    copy_row_dimensions_openpyxl(source_ws, target_ws, 1, copy_end_row)

    target_wb.save(target_file)
    target_wb.close()

    print("  Values/formulas saved.")


# ----------------------------
# Excel COM format/unprotect helpers
# ----------------------------

def unprotect_all_sheets_excel(workbook):
    for ws in workbook.Worksheets:
        try:
            ws.Unprotect(UNPROTECT_PASSWORD)
        except Exception as exc:
            print(f"  WARNING: Could not unprotect worksheet '{ws.Name}' with password: {exc}")


def paste_formats_excel_fresh_copy(excel, source_wb, source_ws, target_wb, target_ws, source_address, target_address):
    """
    V16:
        Re-activates source and target context and copies fresh before each paste.
    """
    source_wb.Activate()
    source_ws.Activate()
    excel.Goto(source_ws.Range(source_address), True)
    source_ws.Range(source_address).Copy()

    target_wb.Activate()
    target_ws.Activate()
    excel.Goto(target_ws.Range(target_address), True)
    target_ws.Range(target_address).PasteSpecial(Paste=XL_PASTE_FORMATS)

    excel.CutCopyMode = False


def paste_column_widths_excel_fresh_copy(excel, source_wb, source_ws, target_wb, target_ws, start_col, end_col):
    """
    V16:
        Re-copy widths for each target workbook.
    """
    source_wb.Activate()
    source_ws.Activate()

    for col in range(start_col, end_col + 1):
        source_ws.Columns(col).Copy()

        target_wb.Activate()
        target_ws.Activate()
        target_ws.Columns(col).PasteSpecial(Paste=XL_PASTE_COLUMN_WIDTHS)

        excel.CutCopyMode = False
        source_wb.Activate()
        source_ws.Activate()


def apply_formats_and_unprotect_with_excel(files_to_process, ppprojects_file, source_copy_settings):
    """
    Uses real Excel to paste formats only.
    V16 re-copies source formatting fresh for each target workbook.
    """
    excel = None
    source_wb = None

    copy_end_col = source_copy_settings["copy_end_col"]
    copy_end_row = source_copy_settings["copy_end_row"]
    copy_end_col_letter = excel_col_letter(copy_end_col)

    source_g_address = f"G1:G{copy_end_row}"
    source_j_end_address = f"J1:{copy_end_col_letter}{copy_end_row}"

    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.EnableEvents = False
        excel.ScreenUpdating = False
        excel.AskToUpdateLinks = False

        source_wb = excel.Workbooks.Open(str(ppprojects_file.resolve()), UpdateLinks=0, ReadOnly=True)
        source_ws = find_excel_sheet_case_insensitive(source_wb, SOURCE_SHEET_NAME)

        if source_ws is None:
            raise ValueError(
                f"{ppprojects_file.name} does not contain a sheet named '{SOURCE_SHEET_NAME}'."
            )

        print("\nApplying Excel Paste Special > Formats:")
        print(f"  Format range: Column G and Columns J:{copy_end_col_letter}")
        print(f"  Format rows: 1:{copy_end_row}")

        for target_file in files_to_process:
            print(f"\nApplying formats/unprotecting with Excel: {target_file.name}")

            target_wb = excel.Workbooks.Open(str(target_file.resolve()), UpdateLinks=0, ReadOnly=False)

            try:
                target_ws = find_excel_sheet_case_insensitive(target_wb, TARGET_SHEET_NAME)

                if target_ws is None:
                    raise ValueError(
                        f"{target_file.name} does not contain a sheet named '{TARGET_SHEET_NAME}'."
                    )

                print("  Unprotecting worksheets...")
                unprotect_all_sheets_excel(target_wb)

                print(f"  Pasting fresh formats for Column G: {source_g_address}")
                paste_formats_excel_fresh_copy(
                    excel=excel,
                    source_wb=source_wb,
                    source_ws=source_ws,
                    target_wb=target_wb,
                    target_ws=target_ws,
                    source_address=source_g_address,
                    target_address=f"G1:G{copy_end_row}"
                )

                print(f"  Pasting fresh formats for Columns J:{copy_end_col_letter}: {source_j_end_address}")
                paste_formats_excel_fresh_copy(
                    excel=excel,
                    source_wb=source_wb,
                    source_ws=source_ws,
                    target_wb=target_wb,
                    target_ws=target_ws,
                    source_address=source_j_end_address,
                    target_address=f"J1:{copy_end_col_letter}{copy_end_row}"
                )

                print("  Pasting fresh column widths...")
                paste_column_widths_excel_fresh_copy(
                    excel=excel,
                    source_wb=source_wb,
                    source_ws=source_ws,
                    target_wb=target_wb,
                    target_ws=target_ws,
                    start_col=SOURCE_COL_G,
                    end_col=SOURCE_COL_G
                )

                paste_column_widths_excel_fresh_copy(
                    excel=excel,
                    source_wb=source_wb,
                    source_ws=source_ws,
                    target_wb=target_wb,
                    target_ws=target_ws,
                    start_col=SOURCE_START_COL_J,
                    end_col=copy_end_col
                )

                excel.CutCopyMode = False
                target_wb.Save()
                print("  Formats saved.")

            finally:
                target_wb.Close(SaveChanges=True)
                excel.CutCopyMode = False

        source_wb.Close(SaveChanges=False)

    finally:
        if excel is not None:
            try:
                excel.CutCopyMode = False
            except Exception:
                pass
            try:
                excel.DisplayAlerts = True
                excel.EnableEvents = True
                excel.ScreenUpdating = True
            except Exception:
                pass
            try:
                excel.Quit()
            except Exception:
                pass


# ----------------------------
# Main
# ----------------------------

def main():
    print("\nPrepare Quarter Workpapers - V16")
    print("Hybrid: openpyxl for data/formulas, Excel fresh PasteSpecial Formats for every workbook\n")

    script_dir = Path(__file__).resolve().parent

    folder_name = input("Enter the workpaper_files folder name located in this script folder: ").strip()
    reporting_period_input = input("Enter the current reporting period, e.g. Q2 2026: ").strip()

    reporting_period = safe_filename_part(reporting_period_input)

    if not folder_name:
        raise ValueError("Folder name cannot be blank.")

    if not reporting_period:
        raise ValueError("Reporting period cannot be blank.")

    workpaper_folder = script_dir / folder_name
    exclusions_path = script_dir / EXCLUSIONS_FILE_NAME

    if not workpaper_folder.exists() or not workpaper_folder.is_dir():
        raise FileNotFoundError(f"Folder not found: {workpaper_folder}")

    if not exclusions_path.exists():
        raise FileNotFoundError(f"Lookup workbook not found: {exclusions_path}")

    lookup_values = read_worksheet_lookup_values(exclusions_path)
    excel_files = get_excel_files(workpaper_folder)
    ppprojects_file = find_ppprojects_file(excel_files)

    print(f"\nWorkpaper folder: {workpaper_folder}")
    print(f"Lookup workbook: {exclusions_path.name}")
    print(f"PPPROJECTS source file: {ppprojects_file.name}")
    print(f"Reporting period: {reporting_period}")

    rename_plan = build_rename_plan(
        excel_files=excel_files,
        ppprojects_file=ppprojects_file,
        lookup_values=lookup_values,
        reporting_period=reporting_period
    )

    if not rename_plan:
        raise ValueError("No workpaper files were matched for rename/update.")

    print("\nRename Preview:")
    files_to_process = []

    for old_path, new_path, matched_value, action in rename_plan:
        if action == "already_named":
            print(f"  ALREADY NAMED | Match: {matched_value} | {old_path.name}")
        else:
            print(f"  MATCH: {matched_value}")
            print(f"    OLD: {old_path.name}")
            print(f"    NEW: {new_path.name}")

    proceed = get_yes_no("\nType YES to rename and update these workpapers: ")

    if not proceed:
        print("\nCancelled. No files were changed.")
        return

    for old_path, new_path, matched_value, action in rename_plan:
        if action == "already_named":
            files_to_process.append(new_path)
            continue

        if new_path.exists():
            raise FileExistsError(
                f"Cannot rename '{old_path.name}' because '{new_path.name}' already exists."
            )

        old_path.rename(new_path)
        files_to_process.append(new_path)

    print("\nRename complete.")

    # Open source once for openpyxl values/formulas.
    source_wb = load_workbook(ppprojects_file)
    source_ws = find_sheet_case_insensitive(source_wb, SOURCE_SHEET_NAME)

    if source_ws is None:
        source_wb.close()
        raise ValueError(
            f"{ppprojects_file.name} does not contain a sheet named '{SOURCE_SHEET_NAME}'."
        )

    source_copy_settings = calculate_source_copy_settings(source_ws)

    print("\nSource copy settings locked:")
    print(f"  Row {ROW_TO_IDENTIFY_END_COLUMN} last used/formatted column: "
          f"{col_letter_openpyxl(source_ws, source_copy_settings['row_6_last_col'])}")
    print(f"  Copy end column including +{EXTRA_COLUMNS_TO_COPY}: "
          f"{col_letter_openpyxl(source_ws, source_copy_settings['copy_end_col'])}")
    print(f"  Copy end row including +{EXTRA_ROWS_TO_COPY}: "
          f"{source_copy_settings['copy_end_row']}")

    # Step 1: copy values/formulas using the stable openpyxl flow.
    for target_file in files_to_process:
        update_target_values_with_openpyxl(target_file, source_ws, source_copy_settings)

    source_wb.close()

    # Step 2: use real Excel only to paste formatting/highlighting and unprotect.
    apply_formats_and_unprotect_with_excel(files_to_process, ppprojects_file, source_copy_settings)

    print("\nAll workpapers updated successfully.")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print("\nERROR:")
        print(exc)
        print("\nScript stopped.")
        input("\nPress Enter to exit...")
        sys.exit(1)
