
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from datetime import datetime, date
from copy import copy

# =========================
# User Prompts
# =========================
source_subfolder = input("📁 Enter the name of the subfolder containing the Excel files: ").strip()
start_date_str   = input("📅 Enter Start Date (MM/DD/YYYY): ").strip()
end_date_str     = input("📅 Enter End Date (MM/DD/YYYY): ").strip()

# =========================
# Parse Dates (as real dates)
# =========================
try:
    start_date = datetime.strptime(start_date_str, "%m/%d/%Y").date()
    end_date   = datetime.strptime(end_date_str, "%m/%d/%Y").date()
except ValueError:
    print("❌ Invalid date format. Please use MM/DD/YYYY.")
    raise SystemExit(1)

# =========================
# Config
# =========================
sheet_name = "Department"
output_sheet_name = "NS Import_Time"

# Header layout (0-based indices; actual Excel rows are +1)
header_row_index           = 5  # actual row 6 (column headers)
projectid_row_index        = 4  # actual row 5
project_stage_row_index    = 3  # actual row 4
project_notes_row_index    = 2  # actual row 3
terminated_dates_row_index = 1  # actual row 2 (project-level termination)
cod_date_row_index         = 0  # actual row 1

# Treat first N columns as metadata (A..F); skipping G and the first project column H
NUM_METADATA_COLS = 6  # A..F; skip G
# Final output column order (STRICT)
FINAL_COLUMNS = [
    "PP Project",
    "ProjectID",
    "Project Stage",
    "Project Notes",
    "Proj Terminated Date",
    "COD Date",
    "Employee",
    "Percent",
    "Emp Id",
    "External ID",
    "Vendor",
    "Contractor",
    "Emp Termination Date",
    "Department",
    "Start Date",
    "End Date",
    "Exclusions",
    "Reason",
]

# =========================
# Paths
# =========================
base_dir = Path(__file__).parent
input_dir = base_dir / source_subfolder
exclusion_file = base_dir / "Exclusions.xlsx"

if not input_dir.exists():
    print(f"❌ Subfolder not found: {input_dir}")
    raise SystemExit(1)
if not exclusion_file.exists():
    print("❌ Exclusions.xlsx not found in script folder.")
    raise SystemExit(1)

# =========================
# Load Exclusion Rules
# =========================
pp_exclude_df = pd.read_excel(exclusion_file, sheet_name="PP Project")
employee_exclude_df = pd.read_excel(exclusion_file, sheet_name="Employee Name")

def _norm_series(s):
    return s.astype(str).str.strip()

pp_exclude_df.columns = [c.strip() for c in pp_exclude_df.columns]
employee_exclude_df.columns = [c.strip() for c in employee_exclude_df.columns]

pp_exclude_map = dict(zip(_norm_series(pp_exclude_df.iloc[:,0]).str.lower(), _norm_series(pp_exclude_df.iloc[:,1]) if pp_exclude_df.shape[1] > 1 else ""))

# Employee exclusions now key off Employee Number from the exclusion tab
employee_number_col = "Employee Number" if "Employee Number" in employee_exclude_df.columns else (employee_exclude_df.columns[2] if employee_exclude_df.shape[1] > 2 else employee_exclude_df.columns[0])
employee_reason_col = "Reason" if "Reason" in employee_exclude_df.columns else (employee_exclude_df.columns[1] if employee_exclude_df.shape[1] > 1 else None)
employee_exclude_map = dict(zip(
    _norm_series(employee_exclude_df[employee_number_col]).str.lower(),
    _norm_series(employee_exclude_df[employee_reason_col]) if employee_reason_col else ""
))

# Load Special Sheets lookup rules
# Expected columns on Exclusions.xlsx > Special Sheets:
# Column A = Sheet Name
# Column B = ProjectID
try:
    special_sheets_df = pd.read_excel(exclusion_file, sheet_name="Special Sheets")
    special_sheets_df.columns = [str(c).strip() for c in special_sheets_df.columns]

    special_sheet_name_col = "Sheet Name" if "Sheet Name" in special_sheets_df.columns else special_sheets_df.columns[0]
    special_projectid_col = "ProjectID" if "ProjectID" in special_sheets_df.columns else (
        "Project ID" if "Project ID" in special_sheets_df.columns else special_sheets_df.columns[1]
    )

    special_sheets_df = special_sheets_df.dropna(subset=[special_sheet_name_col, special_projectid_col])
    special_sheet_map = {}
    for _, special_row in special_sheets_df.iterrows():
        special_project_id = str(special_row[special_projectid_col]).strip().lower()
        special_sheet_name = str(special_row[special_sheet_name_col]).strip()
        if special_project_id and special_sheet_name:
            special_sheet_map.setdefault(special_project_id, []).append(special_sheet_name)

    # Notify when one ProjectID maps to multiple special sheets
    for special_project_id, sheet_names in special_sheet_map.items():
        unique_sheet_names = list(dict.fromkeys(sheet_names))
        special_sheet_map[special_project_id] = unique_sheet_names
        if len(unique_sheet_names) > 1:
            print(f"⚠️ ProjectID '{special_project_id}' maps to multiple Special Sheets: {', '.join(unique_sheet_names)}")

except ValueError:
    print("⚠️ 'Special Sheets' sheet not found in Exclusions.xlsx. Special sheet output will be skipped.")
    special_sheet_map = {}

# =========================
# Helpers
# =========================
def detect_and_scale_percent(s: pd.Series) -> pd.Series:
    # Maintain existing behavior: if any > 1, treat as whole percent
    if (s.dropna() > 1).any():
        s = s / 100.0
    return s

def first_existing(columns, candidates):
    cols = []
    for c in candidates:
        match = [col for col in columns if str(col).strip() == c]
        if match:
            cols.append(match[0])
    return cols

def parse_cell_to_date(val):
    """Robustly parse many date-like inputs to a Python date, else None.
    - Handles datetime/date objects
    - Handles Excel serials (ints/floats) via pandas to_datetime(origin='1899-12-30')
    - Handles strings like '2025-01-05', '01/05/2025', etc.
    - Returns None if not parseable
    """
    import pandas as _pd
    from datetime import date, datetime as _dt
    if val is None or (isinstance(val, float) and _pd.isna(val)):
        return None
    # Already a date/datetime
    if isinstance(val, _dt):
        return val.date()
    if isinstance(val, date):
        return val
    # Excel serial (common case)
    if isinstance(val, (int, float)):
        try:
            ts = _pd.to_datetime(val, unit='D', origin='1899-12-30', errors='coerce')
            return ts.date() if _pd.notna(ts) else None
        except Exception:
            pass
    # Fallback: generic string/obj parse
    try:
        ts = _pd.to_datetime(val, errors='coerce')
        return ts.date() if _pd.notna(ts) else None
    except Exception:
        return None

    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    try:
        if isinstance(val, (int, float)):
            if 1 <= int(val) <= 60000:  # plausible Excel serial
                ts = pd.to_datetime(val, unit="D", origin="1899-12-30", errors="coerce")
                if pd.notna(ts):
                    return ts.date()
    except Exception:
        pass
    try:
        ts = pd.to_datetime(str(val), errors="coerce", infer_datetime_format=True)
        if pd.notna(ts):
            return ts.date()
    except Exception:
        pass
    return None

def safe_excel_sheet_title(raw_title, existing_titles):
    """Return a valid Excel sheet title while staying as close to the lookup value as possible."""
    invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
    title = str(raw_title).strip()
    for ch in invalid_chars:
        title = title.replace(ch, '-')
    if not title:
        title = "Special Sheet"
    title = title[:31]

    # If this title already exists, use it as-is. openpyxl can write to the existing sheet.
    if title in existing_titles:
        return title
    return title

def apply_output_formats(ws, header_names):
    """Apply the same core formatting used by NS Import_Time to an output worksheet."""
    def _col_idx(name):
        return header_names.index(name) + 1 if name in header_names else None

    percent_idx = _col_idx("Percent")
    date_indices = [
        _col_idx("Start Date"),
        _col_idx("End Date"),
        _col_idx("Proj Terminated Date"),
        _col_idx("Emp Termination Date"),
        _col_idx("COD Date"),
    ]

    if percent_idx:
        for rows in ws.iter_rows(min_row=2, min_col=percent_idx, max_col=percent_idx):
            for cell in rows:
                cell.number_format = "0.00%"

    for idx in date_indices:
        if idx:
            for rows in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                for cell in rows:
                    cell.number_format = "MM/DD/YYYY"

    max_width_cap = 80
    for col_cells in ws.columns:
        try:
            col_letter = get_column_letter(col_cells[0].column)
        except Exception:
            continue
        max_len = 0
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width_cap)

    ws.auto_filter.ref = ws.dimensions



def move_yellow_highlighted_rows_to_bottom(ws):
    """Move rows with a standard yellow fill to the bottom of NS Import_Time.
    Preserves the existing row order within the clear and highlighted groups.
    """
    if ws.max_row <= 2:
        return 0

    def _is_standard_yellow(cell):
        fill = cell.fill
        if not fill or fill.fill_type != "solid":
            return False
        rgb = getattr(fill.fgColor, "rgb", None)
        # openpyxl may store FFFF00 as 00FFFF00 or FFFFFF00 depending on how the fill was created.
        return bool(rgb and str(rgb).upper().endswith("FFFF00"))

    row_packages = []
    yellow_count = 0

    for row_num in range(2, ws.max_row + 1):
        cells = []
        row_is_yellow = False
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_num, column=col_num)
            if _is_standard_yellow(cell):
                row_is_yellow = True
            cells.append({
                "value": cell.value,
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "fill": copy(cell.fill),
                "font": copy(cell.font),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            })
        if row_is_yellow:
            yellow_count += 1
        row_packages.append((row_is_yellow, cells))

    # Stable partition: clear rows first, yellow-highlighted rows last.
    row_packages.sort(key=lambda item: item[0])

    for target_row_num, (_, cells) in enumerate(row_packages, start=2):
        for col_num, cell_info in enumerate(cells, start=1):
            cell = ws.cell(row=target_row_num, column=col_num)
            cell.value = cell_info["value"]
            cell._style = copy(cell_info["style"])
            cell.number_format = cell_info["number_format"]
            cell.fill = copy(cell_info["fill"])
            cell.font = copy(cell_info["font"])
            cell.border = copy(cell_info["border"])
            cell.alignment = copy(cell_info["alignment"])
            cell.protection = copy(cell_info["protection"])

    return yellow_count

# =========================
# Processing
# =========================
excel_files = sorted([p for p in input_dir.glob("*.xlsx") if p.name.lower() != "exclusions.xlsx"])

if not excel_files:
    print("Warning: No .xlsx files found to process.")
    raise SystemExit(0)

for fpath in excel_files:
    print(f"➡️ Processing: {fpath.name}")

    # Read WITHOUT headers to preserve top bands
    raw = pd.read_excel(fpath, sheet_name=sheet_name, header=None, engine="openpyxl")
    nrows, ncols = raw.shape

    if nrows < header_row_index + 1:
        print(f"   ❌ Not enough rows to build headers (need row {header_row_index+1}). Skipping.")
        continue

    # Build column headers from row 6 (keep raw for blank detection; still use ffill for column naming)
    header_row = raw.iloc[header_row_index]
    header_filled = header_row.astype(str).replace({'': None}).ffill().fillna('Unnamed')
    header_raw_stripped = header_row.apply(lambda v: (str(v).strip() if pd.notna(v) else ''))

    # Data starts row 7
    data = raw.iloc[header_row_index+1:].copy()
    data.columns = header_filled

    if data.shape[1] < NUM_METADATA_COLS:
        print(f"   ❌ Fewer than {NUM_METADATA_COLS} columns; cannot split metadata/project columns reliably. Skipping.")
        continue

    # Split metadata vs project columns by POSITION (A..F metadata, H.. projects)
    meta_cols = list(data.columns[:NUM_METADATA_COLS])  # A..G
    proj_start = NUM_METADATA_COLS + 2  # skip G and H
    proj_end = data.shape[1]

    # Build top-band row Series for positional lookup
    proj_id_vals     = raw.iloc[projectid_row_index]
    proj_stage_vals  = raw.iloc[project_stage_row_index]
    proj_notes_vals  = raw.iloc[project_notes_row_index]
    proj_term_vals   = raw.iloc[terminated_dates_row_index]
    cod_vals         = raw.iloc[cod_date_row_index]
    # Prepare container for long rows
    long_frames = []

    # Loop through project columns by POSITION to preserve duplicates and correct band mapping
    # Guard: ensure we have columns beyond the skipped G and H
    if proj_start >= proj_end:
        print("   Warning: No project columns beyond I after skipping G and H; skipping file.")
        continue
    for colpos in range(proj_start, proj_end):
        proj_name = str(header_filled.iloc[colpos])  # keep original project header name
        percent_series = pd.to_numeric(data.iloc[:, colpos], errors="coerce")
        # Skip this project column if its original row-6 header was blank
        if header_raw_stripped.iloc[colpos] == '':
            continue

        # Build a mini-frame for this project
        # Build a mini-frame for this project
        mini = pd.DataFrame(index=data.index)
        # Carry ALL metadata columns A..G exactly as-is (simple copy/paste of values)
        try:
            mini[meta_cols] = data[meta_cols].copy()
        except Exception:
            # Fallback: add columns one-by-one if shapes mismatch
            for mcol in meta_cols:
                mini[mcol] = data[mcol] if mcol in data.columns else None
        # Attach known fields
        mini["PP Project"] = proj_name
        mini["Percent"] = percent_series

        # Project-top-band fields by POSITION
        def _val(series):  # helper captures position and convert where needed
            return series.iloc[colpos] if colpos < len(series) else None

        mini["ProjectID"]             = _val(proj_id_vals)
        mini["Project Stage"]         = _val(proj_stage_vals)
        mini["Project Notes"]         = _val(proj_notes_vals)
        mini["Proj Terminated Date"]  = parse_cell_to_date(_val(proj_term_vals))
        mini["COD Date"]              = parse_cell_to_date(_val(cod_vals))
        # Only append if this project column has any non-NA Percent (avoids empty/all-NA frames)
        if mini['Percent'].notna().any():
            long_frames.append(mini)

    # Concatenate all project frames (rows stack; duplicate project names allowed)
    # Guard: if no project columns produced data, skip this file
    if not long_frames:
        print("   Warning: No valid project columns with data; skipping file.")
        continue
    long = pd.concat(long_frames, axis=0, ignore_index=True)
# Standardize names for people/id
    rename_map = {
        "Employee Name": "Employee",
        "Employee number": "Emp Id",
    }
    long = long.rename(columns=rename_map)

    # External ID = Emp Id when present
    if "Emp Id" in long.columns:
        long["External ID"] = long["Emp Id"]
    else:
        long["External ID"] = None

    # Add Start/End as real dates
    long["Start Date"] = start_date
    long["End Date"]   = end_date

    # Percent clean
    long = long[long["Percent"].notna() & (long["Percent"] != 0)]
    long["Percent"] = detect_and_scale_percent(long["Percent"])

    # Emp Termination Date (from metadata "Termination Date" if present)
    if "Termination Date" in long.columns:
        long["Emp Termination Date"] = long["Termination Date"].apply(parse_cell_to_date)
        long = long.drop(columns=["Termination Date"], errors="ignore")

    # =========================
    # Apply Exclusions (order: Notes -> Project -> Employee)
    # =========================
    long["Exclusions"] = "Include"
    long["Reason"] = ""

    # 1) Project Notes present -> Exclude
    if "Project Notes" in long.columns:
        mask_notes = long["Project Notes"].notna() & (long["Project Notes"].astype(str).str.strip() != "")
        long.loc[mask_notes, "Exclusions"] = "Exclude"
        long.loc[mask_notes, "Reason"] = long.loc[mask_notes, "Project Notes"].astype(str).str.strip()

    # 2) PP Project map (case-insensitive match), only if not already excluded
    if "PP Project" in long.columns:
        proj_key = long["PP Project"].astype(str).str.strip().str.lower()
        mask_proj = (long["Exclusions"] != "Exclude") & proj_key.isin(pp_exclude_map.keys())
        long.loc[mask_proj, "Exclusions"] = "Exclude"
        long.loc[mask_proj, "Reason"] = proj_key.map(pp_exclude_map)

    # 3) Employee Number map (case-insensitive), only if not already excluded
    if "Emp Id" in long.columns:
        emp_key = long["Emp Id"].astype(str).str.strip().str.lower()
        mask_emp = (long["Exclusions"] != "Exclude") & emp_key.isin(employee_exclude_map.keys())
        long.loc[mask_emp, "Exclusions"] = "Exclude"
        long.loc[mask_emp, "Reason"] = emp_key.map(employee_exclude_map)

    # =========================
    # STRICT final column order
    # =========================
    for col in FINAL_COLUMNS:
        if col not in long.columns:
            long[col] = pd.NA
    long = long[FINAL_COLUMNS]

    # =========================
    # Sorting: Exclusions (Include first), then Employee A–Z, then PP Project A–Z
    # =========================
    exc_norm = long["Exclusions"].astype(str).str.strip().str.lower()
    long["_exc_rank"] = (exc_norm == "include").astype(int)  # Include=1, Exclude=0
    long = long.sort_values(["_exc_rank", "Employee", "PP Project"],
                            ascending=[False, True, True],
                            na_position="last").drop(columns=["_exc_rank"])

    # =========================
    # Write back to workbook (create/replace NS Import_Time)
    # =========================
    wb = load_workbook(fpath)
    if output_sheet_name in wb.sheetnames:
        ws_old = wb[output_sheet_name]
        wb.remove(ws_old)
    ws = wb.create_sheet(output_sheet_name)

    # Write data
    # --- Ensure no pandas.NA reach openpyxl ---
    import numpy as _np
    long = long.where(long.notna(), None)  # replace pd.NA/NaN with None for Excel compatibility
    for row in dataframe_to_rows(long, index=False, header=True):
        ws.append([None if (getattr(v, 'to_pydatetime', None) is None and hasattr(v, '__class__') and str(v) == '<NA>') or v is pd.NA else v for v in row])

    # --- Rebuild sheet to enforce final column order at the Excel level ---
    desired_headers = FINAL_COLUMNS.copy()

    # Map current header -> column index
    current_headers = [cell.value for cell in ws[1]]
    header_to_idx = {str(h): i+1 for i, h in enumerate(current_headers)}

    # Create a new ordered sheet
    ordered_title = output_sheet_name + "_ordered"
    if ordered_title in wb.sheetnames:
        wb.remove(wb[ordered_title])
    ws2 = wb.create_sheet(ordered_title)

    # Write headers in desired order
    for col_idx, h in enumerate(desired_headers, start=1):
        ws2.cell(row=1, column=col_idx, value=h)

    # Copy row values in desired header order
    max_row = ws.max_row
    for r in range(2, max_row + 1):
        for col_idx, h in enumerate(desired_headers, start=1):
            src_idx = header_to_idx.get(h)
            if src_idx:
                ws2.cell(row=r, column=col_idx, value=ws.cell(row=r, column=src_idx).value)

    # Delete original and rename ordered to final name
    wb.remove(ws)
    ws = ws2
    ws.title = output_sheet_name

    # ---- Sort rows again post-rebuild (same order) & set AutoFilter ----
    header_row = [cell.value for cell in ws[1]]
    def _idx(col_name):
        return header_row.index(col_name) + 1 if col_name in header_row else None

    pp_idx = _idx("PP Project")
    emp_idx = _idx("Employee")
    exc_idx = _idx("Exclusions")

    # Extract data rows
    data_rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        data_rows.append(row_vals)

    def _norm_exc(v):
        s = (str(v).strip().lower() if v is not None else "")
        return 0 if s == "include" else 1  # Include first

    def sort_key(row):
        exc_rank = _norm_exc(row[exc_idx-1]) if exc_idx else 0
        emp = row[emp_idx-1] if emp_idx else ""
        pp = row[pp_idx-1] if pp_idx else ""
        return (exc_rank, str(emp), str(pp))

    data_rows.sort(key=sort_key)

    # Write back sorted rows
    for i, row_vals in enumerate(data_rows, start=2):
        for c, val in enumerate(row_vals, start=1):
            ws.cell(row=i, column=c, value=val)
# Apply number formats
    header_list = [cell.value for cell in ws[1]]
    def col_idx(name):
        return header_list.index(name) + 1 if name in header_list else None

    percent_idx = col_idx("Percent")
    start_idx   = col_idx("Start Date")
    end_idx     = col_idx("End Date")
    proj_term_idx = col_idx("Proj Terminated Date")
    emp_term_idx  = col_idx("Emp Termination Date")
    cod_idx     = col_idx("COD Date")

    # Percent as percent
    if percent_idx:
        for r in ws.iter_rows(min_row=2, min_col=percent_idx, max_col=percent_idx):
            for cell in r:
                cell.number_format = "0.00%"

    # Dates as short date MM/DD/YYYY
    def format_date_col(idx):
        if idx:
            for r in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                for cell in r:
                    cell.number_format = "MM/DD/YYYY"

    for idx in [start_idx, end_idx, proj_term_idx, emp_term_idx, cod_idx]:
        format_date_col(idx)

    # Auto-fit column widths
    max_width_cap = 80
    for col_cells in ws.columns:
        try:
            col_letter = get_column_letter(col_cells[0].column)
        except Exception:
            try:
                col_letter = get_column_letter(col_cells[0].col_idx)
            except Exception:
                continue
        max_len = 0
        for cell in col_cells:
            val = cell.value
            if val is not None:
                val_len = len(str(val))
                if val_len > max_len:
                    max_len = val_len
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width_cap)

    # Add AutoFilter across header
    ws.auto_filter.ref = ws.dimensions

    # =========================
    # Special Sheets Output
    # =========================
    # If NS Import_Time ProjectID matches Exclusions.xlsx > Special Sheets,
    # highlight the NS Import_Time row yellow and append the row to the named special sheet.
    if special_sheet_map:
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ns_headers = [cell.value for cell in ws[1]]
        projectid_idx = ns_headers.index("ProjectID") + 1 if "ProjectID" in ns_headers else None

        if not projectid_idx:
            print("   ⚠️ ProjectID column not found in NS Import_Time. Special Sheets step skipped.")
        else:
            special_rows_added = {}
            for row_num in range(2, ws.max_row + 1):
                ns_project_id_raw = ws.cell(row=row_num, column=projectid_idx).value
                ns_project_id = str(ns_project_id_raw).strip().lower() if ns_project_id_raw is not None else ""

                if ns_project_id in special_sheet_map:
                    row_values = [ws.cell(row=row_num, column=col_num).value for col_num in range(1, ws.max_column + 1)]

                    # Highlight the entire matching NS Import_Time row yellow
                    for col_num in range(1, ws.max_column + 1):
                        ws.cell(row=row_num, column=col_num).fill = yellow_fill

                    for lookup_sheet_name in special_sheet_map[ns_project_id]:
                        special_title = safe_excel_sheet_title(lookup_sheet_name, wb.sheetnames)
                        if special_title != lookup_sheet_name:
                            print(f"   ⚠️ Special sheet name '{lookup_sheet_name}' was adjusted to '{special_title}' for Excel compatibility.")

                        if special_title not in wb.sheetnames:
                            special_ws = wb.create_sheet(special_title)
                            for col_num, header in enumerate(ns_headers, start=1):
                                special_ws.cell(row=1, column=col_num, value=header)
                        else:
                            special_ws = wb[special_title]
                            # If the sheet exists but has no usable headers, add NS Import_Time headers to row 1
                            existing_headers = [special_ws.cell(row=1, column=col_num).value for col_num in range(1, len(ns_headers) + 1)]
                            if all(h is None for h in existing_headers):
                                for col_num, header in enumerate(ns_headers, start=1):
                                    special_ws.cell(row=1, column=col_num, value=header)

                        append_row_num = special_ws.max_row + 1
                        for col_num, value in enumerate(row_values, start=1):
                            special_ws.cell(row=append_row_num, column=col_num, value=value)

                        special_rows_added[special_title] = special_rows_added.get(special_title, 0) + 1

            for special_title, added_count in special_rows_added.items():
                apply_output_formats(wb[special_title], ns_headers)
                print(f"   ✅ Added {added_count} row(s) to Special Sheet '{special_title}'.")

            if not special_rows_added:
                print("   ℹ️ No Special Sheets matches found for this workbook.")

    yellow_row_count = move_yellow_highlighted_rows_to_bottom(ws)
    if yellow_row_count:
        print(f"   ✅ Sorted NS Import_Time so {yellow_row_count} yellow highlighted row(s) are at the bottom.")

    wb.save(fpath)
    print(f"   ✅ Wrote '{output_sheet_name}' in {fpath.name}")

print("🎉 All files processed.")