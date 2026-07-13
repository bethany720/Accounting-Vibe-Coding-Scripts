#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Payroll JE Builder — v29
- Renames the ORIGINAL GL workbook file itself (in place) to:
    Payroll WS_<Bank Date MM.DD.YYYY>_<Employee Paid Date MM.DD.YYYY>.xlsx
  using os.replace (overwrite if exists).
- Worksheet tab remains "Payroll JE".
- CSV name remains:
    Payroll Import_<Bank Date MM.DD.YYYY>_<Employee Paid Date MM.DD.YYYY>.csv
- All other behavior from v28 retained.
"""

from __future__ import annotations

import sys, glob, os, re
from pathlib import Path
from datetime import datetime

import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, numbers

# ---------------- Constants ----------------
ORIG_SHEET_NAME = "CR General Ledger Outbound Matr"
STOP_LABEL = "10012.01:JP Morgan"
EXCLUDE_LABEL = "20600.07:Payroll Suspense"

LOOKUP_BASENAME = "Lookup.xlsx"
OUTPUT_SHEET_BASE = "Payroll JE"  # worksheet/tab name

# Static JE fields (numeric for Excel correctness)
SUBSIDIARY_CONST = 69
MAIN_ACCOUNT = 10012.01
PAYEE = "ONESOURCE VIRTUAL, INC."
CURRENCY = "USD"
DEPARTMENT = "Operating : Current"

IMPORT_COLS = [
    "EXTERNAL ID","DATE","SUBSIDIARY","ACCOUNT (MAIN)","PAYEE","CURRENCY","MEMO (MAIN)",
    "ACCOUNT","AMOUNT","MEMO (LINE)","DEPARTMENT","PP Project"
]

WARNING_MEMO = "WARNING: ACCT NOT IN LOOKUP"

# --------------- File location helpers -------------------
def find_lookup_here(base_dir: Path) -> Path | None:
    p = base_dir / LOOKUP_BASENAME
    return p if p.exists() else None

def find_gl_here(base_dir: Path) -> Path | None:
    patterns = ["*.xlsx","*.xlsm","*.xls"]
    files = []
    for pat in patterns:
        files += [Path(x) for x in glob.glob(str(base_dir / pat))]
    files = [f for f in files if f.name.lower() != LOOKUP_BASENAME.lower() and not f.name.startswith("~$")]
    if not files:
        return None
    if len(files) > 1:
        print("⚠️  Multiple GL workbooks found in the script folder:")
        for f in files:
            print(f"   - %s" % f.name)
        print("Please keep only ONE GL workbook in this folder and re-run.")
        sys.exit(2)
    return files[0]

def ensure_source_sheet(xlsx_path: Path) -> str:
    wb = load_workbook(str(xlsx_path))
    if ORIG_SHEET_NAME not in wb.sheetnames:
        first = wb.sheetnames[0]
        ws = wb[first]
        ws.title = ORIG_SHEET_NAME
        wb.save(str(xlsx_path))
    wb.close()
    return ORIG_SHEET_NAME

def _acct_key(val) -> str:
    try:
        return f"{float(str(val).strip()):.2f}"
    except Exception:
        return str(val).strip()

# --------------- Load Lookup ----------------
def load_lookup(base_dir: Path):
    path = find_lookup_here(base_dir)
    if not path:
        print("!! Lookup.xlsx not found in the script folder. Place it next to this .py file.")
        sys.exit(2)
    df = pd.read_excel(path, sheet_name=0, header=0)
    cols = list(df.columns)
    ren = {}
    if len(cols)>=1: ren[cols[0]] = "A"
    if len(cols)>=2: ren[cols[1]] = "B"
    if len(cols)>=3: ren[cols[2]] = "C"
    if len(cols)>=4: ren[cols[3]] = "D"
    df = df.rename(columns=ren)

    special_map: dict[str, dict[str, tuple[str,str]]] = {}
    non_special_defaults: dict[str, tuple[str,str]] = {}
    seen_pairs = set()

    for _, r in df.iterrows():
        acct = _acct_key(r.get("A",""))
        if not acct:
            continue
        memo = "" if pd.isna(r.get("B","")) else str(r.get("B","")).strip()
        raw_c = r.get("C", 1)
        try:
            s_val = float(str(raw_c).strip())
        except Exception:
            s_val = 1.0
        sign_str = str(int(s_val)) if float(s_val).is_integer() else str(s_val)
        dval = "" if pd.isna(r.get("D","")) else str(r.get("D","")).strip()

        if acct not in non_special_defaults:
            non_special_defaults[acct] = (memo, sign_str)
        else:
            old_memo, old_sign = non_special_defaults[acct]
            if not old_memo and memo:
                old_memo = memo
            non_special_defaults[acct] = (old_memo, old_sign)

        if dval:
            key = (acct, dval)
            if key in seen_pairs:
                raise ValueError("Duplicate data in Lookup")
            seen_pairs.add(key)
            special_map.setdefault(acct, {})[dval] = (memo, sign_str)

    return special_map, non_special_defaults

# --------------- Read GL ----------------
def read_gl(xlsx_path: Path, sheet_name: str):
    try:
        df = pd.read_excel(str(xlsx_path), sheet_name=sheet_name, header=1, usecols=[0,1,2])
    except Exception:
        df = pd.read_excel(str(xlsx_path), sheet_name=sheet_name, header=0, usecols=[0,1,2])

    df = df.rename(columns={df.columns[0]:"A", df.columns[1]:"B", df.columns[2]:"C"})
    df = df[~df["A"].isna()]
    df["A"] = df["A"].astype(str)

    stop_mask = df["A"].str.strip() == STOP_LABEL
    bank_line = df[stop_mask]
    bank_amt = float(pd.to_numeric(bank_line["C"], errors="coerce").fillna(0.0).sum()) if not bank_line.empty else 0.0

    if not bank_line.empty:
        first_stop_idx = bank_line.index[0]
        df = df.loc[:first_stop_idx-1, :]

    df = df[df["A"].str.strip() != EXCLUDE_LABEL]

    def split_prefix(s: str) -> str:
        s = str(s)
        if ":" in s:
            left,_ = s.split(":",1)
            return left.strip()
        return s[:8].strip()

    prefixes = df["A"].map(split_prefix).astype(str)
    work = pd.DataFrame({
        "prefix": prefixes,
        "bval": df["B"].astype(str),
        "amt": pd.to_numeric(df["C"], errors="coerce").fillna(0.0)
    })

    return work, bank_amt

# --------------- Build JE rows ----------------
def build_je_rows(work: pd.DataFrame,
                  paid_dt: datetime,
                  bank_dt: datetime,
                  special_map: dict,
                  non_special_defaults: dict):
    paid_str = paid_dt.strftime("%m/%d/%Y")
    external_id = "CHCK" + paid_dt.strftime("%m%d%y")
    main_memo = f"{paid_str} Payroll"

    rows = []
    per_line = []

    for acct in sorted(work["prefix"].unique()):
        acct_key = _acct_key(acct)
        gl_subset = work[work["prefix"] == acct_key]

        is_special = acct_key in special_map and len(special_map[acct_key]) > 0
        if is_special:
            d_map = special_map[acct_key]
            d_list = sorted(d_map.keys())

            for d in d_list:
                gl_match = gl_subset[gl_subset["bval"].astype(str).str.strip() == d]
                if gl_match.empty:
                    continue
                memo, sign = d_map[d]
                memo_eff = memo if memo else WARNING_MEMO
                line_memo = f"{paid_str} {memo_eff}"
                rows.append({
                    "EXTERNAL ID": external_id,
                    "DATE": bank_dt,
                    "SUBSIDIARY": SUBSIDIARY_CONST,
                    "ACCOUNT (MAIN)": MAIN_ACCOUNT,
                    "PAYEE": PAYEE,
                    "CURRENCY": CURRENCY,
                    "MEMO (MAIN)": main_memo,
                    "ACCOUNT": float(acct_key) if re.match(r'^\d+(\.\d+)?$', acct_key) else acct_key,
                    "AMOUNT": None,
                    "MEMO (LINE)": line_memo,
                    "DEPARTMENT": DEPARTMENT,
                    "PP Project": ""
                })
                per_line.append({
                    "acct": acct_key, "mode":"d_split", "d_list": d_list, "d_value": d,
                    "sign": sign, "is_catchall": False
                })

            d_set = set(d_list)
            leftovers = gl_subset[~gl_subset["bval"].astype(str).str.strip().isin(d_set)]
            if not leftovers.empty:
                default_sign = non_special_defaults.get(acct_key, ("","1"))[1]
                line_memo = f"{paid_str} {WARNING_MEMO}"
                rows.append({
                    "EXTERNAL ID": external_id,
                    "DATE": bank_dt,
                    "SUBSIDIARY": SUBSIDIARY_CONST,
                    "ACCOUNT (MAIN)": MAIN_ACCOUNT,
                    "PAYEE": PAYEE,
                    "CURRENCY": CURRENCY,
                    "MEMO (MAIN)": main_memo,
                    "ACCOUNT": float(acct_key) if re.match(r'^\d+(\.\d+)?$', acct_key) else acct_key,
                    "AMOUNT": None,
                    "MEMO (LINE)": line_memo,
                    "DEPARTMENT": DEPARTMENT,
                    "PP Project": ""
                })
                per_line.append({
                    "acct": acct_key, "mode":"catchall", "d_list": d_list, "d_value": None,
                    "sign": default_sign, "is_catchall": True
                })

        else:
            default_memo, default_sign = non_special_defaults.get(acct_key, ("", "1"))
            memo_eff = default_memo if default_memo else WARNING_MEMO
            line_memo = f"{paid_str} {memo_eff}"
            rows.append({
                "EXTERNAL ID": external_id,
                "DATE": bank_dt,
                "SUBSIDIARY": SUBSIDIARY_CONST,
                "ACCOUNT (MAIN)": MAIN_ACCOUNT,
                "PAYEE": PAYEE,
                "CURRENCY": CURRENCY,
                "MEMO (MAIN)": main_memo,
                "ACCOUNT": float(acct_key) if re.match(r'^\d+(\.\d+)?$', acct_key) else acct_key,
                "AMOUNT": None,
                "MEMO (LINE)": line_memo,
                "DEPARTMENT": DEPARTMENT,
                "PP Project": ""
            })
            per_line.append({
                "acct": acct_key, "mode":"normal", "d_list": [], "d_value": None,
                "sign": default_sign, "is_catchall": False
            })

    je_df = pd.DataFrame(rows, columns=IMPORT_COLS)

    # Ensure numeric types
    je_df["SUBSIDIARY"] = pd.to_numeric(je_df["SUBSIDIARY"], errors="coerce")
    je_df["ACCOUNT (MAIN)"] = pd.to_numeric(je_df["ACCOUNT (MAIN)"], errors="coerce")

    return je_df, per_line

# --------------- Write Excel with formulas ----------------
def write_je_with_formulas(xlsx_path: Path, je_df: pd.DataFrame, per_line: list):
    wb = load_workbook(str(xlsx_path))

    base = OUTPUT_SHEET_BASE
    sheet_name = base
    i = 1
    while sheet_name in wb.sheetnames:
        i += 1
        sheet_name = f"{base} {i}"
    ws = wb.create_sheet(title=sheet_name)

    ws.append(IMPORT_COLS)
    for _, row in je_df.iterrows():
        ws.append(list(row.values))

    bold = Font(bold=True)
    for c in range(1, len(IMPORT_COLS)+1):
        ws.cell(row=1, column=c).font = bold
    widths = [20,12,12,14,28,8,22,12,16,40,22,12]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    for r in range(2, ws.max_row+1):
        ws.cell(row=r, column=2).number_format = "mm/dd/yyyy"        # DATE
        ws.cell(row=r, column=3).number_format = "0"                 # SUBSIDIARY
        ws.cell(row=r, column=4).number_format = "0.00"              # ACCOUNT (MAIN)
        ws.cell(row=r, column=8).number_format = "0.00"              # ACCOUNT
        ws.cell(row=r, column=9).number_format = numbers.FORMAT_NUMBER_00  # AMOUNT

    # Formulas in AMOUNT
    for idx, r in enumerate(range(2, ws.max_row+1), start=0):
        meta = per_line[idx]
        acct_ref = f"H{r}"
        sign = meta["sign"]
        try:
            fval = float(str(sign).strip())
            sign_lit = str(int(fval)) if fval.is_integer() else str(fval)
        except Exception:
            sign_lit = "1"

        if meta["mode"] == "d_split":
            d = meta["d_value"].replace('"','""')
            core = (
                f"SUMIFS('{ORIG_SHEET_NAME}'!$C:$C,"
                f"'{ORIG_SHEET_NAME}'!$A:$A,{acct_ref}&\"*\","
                f"'{ORIG_SHEET_NAME}'!$B:$B,\"{d}\")"
            )
            formula = f"=IFERROR(({core})*{sign_lit},0)"
        elif meta["mode"] == "catchall":
            parts = [f"SUMIF('{ORIG_SHEET_NAME}'!$A:$A,{acct_ref}&\"*\",'{ORIG_SHEET_NAME}'!$C:$C)"]
            for d in meta["d_list"]:
                d_esc = d.replace('"','""')
                parts.append(
                    f"-SUMIFS('{ORIG_SHEET_NAME}'!$C:$C,"
                    f"'{ORIG_SHEET_NAME}'!$A:$A,{acct_ref}&\"*\","
                    f"'{ORIG_SHEET_NAME}'!$B:$B,\"{d_esc}\")"
                )
            core = "(" + "".join(parts) + ")"
            formula = f"=IFERROR(({core})*{sign_lit},0)"
        else:  # normal
            core = (
                f"SUMIF('{ORIG_SHEET_NAME}'!$A:$A,{acct_ref}&\"*\",'{ORIG_SHEET_NAME}'!$C:$C)"
            )
            formula = f"=IFERROR(({core})*{sign_lit},0)"

        ws.cell(row=r, column=9, value=formula)

    wb.save(str(xlsx_path))
    wb.close()
    return sheet_name

# --------------- Python recompute for CSV ----------------
def recompute_amounts(work: pd.DataFrame, je_df: pd.DataFrame, per_line: list) -> pd.Series:
    prefixes = work["prefix"].astype(str).values
    bvals = work["bval"].astype(str).values
    amts = work["amt"].astype(float).values

    out = []
    for idx, row in je_df.iterrows():
        meta = per_line[idx]
        acct = str(row["ACCOUNT"]).strip()
        try:
            acct = f"{float(acct):.2f}"
        except Exception:
            pass
        try:
            mult = float(str(meta["sign"]).strip())
        except Exception:
            mult = 1.0

        if meta["mode"] == "d_split":
            d = meta["d_value"]
            total = 0.0
            for p, b, a in zip(prefixes, bvals, amts):
                if p == acct and str(b).strip() == d:
                    total += float(a)
            out.append(mult * total)

        elif meta["mode"] == "catchall":
            d_set = set(meta["d_list"])
            total = 0.0
            for p, b, a in zip(prefixes, bvals, amts):
                if p == acct and str(b).strip() not in d_set:
                    total += float(a)
            out.append(mult * total)

        else:  # normal
            total = 0.0
            for p, a in zip(prefixes, amts):
                if p == acct:
                    total += float(a)
            out.append(mult * total)

    return pd.Series(out, index=je_df.index)

# --------------- CSV export ----------------
def export_csv(out_dir: Path, bank_dt: datetime, paid_dt: datetime, je_df: pd.DataFrame) -> Path:
    out = je_df.copy()
    out["DATE"] = out["DATE"].dt.strftime("%m/%d/%Y")
    out["ACCOUNT (MAIN)"] = pd.to_numeric(out["ACCOUNT (MAIN)"], errors="coerce")
    out["ACCOUNT"] = pd.to_numeric(out["ACCOUNT"], errors="coerce")
    out["SUBSIDIARY"] = pd.to_numeric(out["SUBSIDIARY"], errors="coerce")
    out["AMOUNT"] = pd.to_numeric(out["AMOUNT"], errors="coerce").round(2).fillna(0.0)

    missing = [c for c in IMPORT_COLS if c not in out.columns]
    if missing:
        raise KeyError(f"Missing expected columns: {missing}")
    out = out[IMPORT_COLS]

    bank_lbl = bank_dt.strftime("%m.%d.%Y")
    paid_lbl = paid_dt.strftime("%m.%d.%Y")
    out_name = f"Payroll Import_{bank_lbl}_{paid_lbl}.csv"
    out_path = out_dir / out_name
    out.to_csv(out_path, index=False, float_format="%.2f", encoding="utf-8-sig")
    return out_path

# --------------- Warning summary ----------------
def print_warning_summary(je_df: pd.DataFrame):
    if "MEMO (LINE)" not in je_df.columns:
        return
    warn_mask = je_df["MEMO (LINE)"].astype(str).str.contains("WARNING: ACCT NOT IN LOOKUP", na=False)
    warn_rows = je_df[warn_mask]
    if warn_rows.empty:
        print("   No WARNING lines were generated. ✅")
        return

    accts = warn_rows["ACCOUNT"].astype(str).tolist()
    def _norm(a):
        try:
            return f"{float(a):.2f}"
        except Exception:
            return a
    unique_accts = sorted({_norm(a) for a in accts})
    print(f"\n⚠️  WARNING: ACCT NOT IN LOOKUP used on {len(warn_rows)} line(s).")
    print(f"   Accounts involved ({len(unique_accts)}):")
    for a in unique_accts[:1000]:
        print(f"     - {a}")

# --------------- Main ----------------
def main():
    base_dir = Path(__file__).resolve().parent

    # Locate Lookup
    lookup_path = find_lookup_here(base_dir)
    if not lookup_path:
        print("!! Lookup.xlsx not found in the script folder. Place it next to this .py file.")
        sys.exit(2)

    # Locate GL workbook (only one allowed)
    gl_path = find_gl_here(base_dir)
    if not gl_path:
        print("!! No GL workbook (*.xlsx / *.xlsm / *.xls) found in the script folder.")
        sys.exit(2)

    try:
        bank_str = input("Bank Date for the JE (MM/DD/YYYY): ").strip()
        paid_str = input("Employee Paid Date (MM/DD/YYYY): ").strip()
        bank_dt = datetime.strptime(bank_str, "%m/%d/%Y")
        paid_dt = datetime.strptime(paid_str, "%m/%d/%Y")
    except Exception:
        print("!! Please enter dates as MM/DD/YYYY."); sys.exit(2)

    ensure_source_sheet(gl_path)
    try:
        special_map, non_special_defaults = load_lookup(base_dir)
    except FileNotFoundError as e:
        print(str(e)); sys.exit(2)
    except ValueError as e:
        print(str(e)); sys.exit(2)

    work_df, bank_amt = read_gl(gl_path, ORIG_SHEET_NAME)

    je_df, per_line = build_je_rows(work_df, paid_dt, bank_dt, special_map, non_special_defaults)

    # Write JE + formulas into the GL workbook (tab stays "Payroll JE")
    sheet_name = write_je_with_formulas(gl_path, je_df, per_line)

    # Python recompute for CSV
    py_amounts = recompute_amounts(work_df, je_df, per_line).round(2)
    je_df = je_df.copy()
    je_df["AMOUNT"] = py_amounts

    # Export CSV to current folder with dot-date naming
    out_csv = export_csv(base_dir, bank_dt, paid_dt, je_df)

    # ---- Final Checker (flip total) ----
    total_je_raw = float(py_amounts.sum())
    total_je = total_je_raw * -1.0  # flip sign for comparison to positive bank debit
    ok = abs(total_je - bank_amt) < 0.005

    print("\n---- RESULT ----")
    print(f"   GL workbook: {gl_path.name}")
    print(f"   New sheet:   {sheet_name}")
    print(f"   CSV written: {out_csv.name}")
    print(f"   JE total (flipped): {total_je:,.2f}")
    print(f"   Bank debit:         {bank_amt:,.2f}")
    if not ok:
        print(f"   (Raw JE total before flip: {total_je_raw:,.2f})")
    print("   ✅ Totals match the bank debit line." if ok else "   ⚠️  Totals DO NOT match.")

    # Warning summary
    print_warning_summary(je_df)

    # ---- Rename the ORIGINAL GL workbook in place ----
    bank_lbl = bank_dt.strftime("%m.%d.%Y")
    paid_lbl = paid_dt.strftime("%m.%d.%Y")
    new_name = f"Payroll WS_{bank_lbl}_{paid_lbl}.xlsx"
    dest = base_dir / new_name
    try:
        os.replace(str(gl_path), str(dest))  # atomic replace; overwrites if exists
        print(f"\n✏️  Renamed GL workbook to: {new_name}")
    except Exception as e:
        print(f"\n⚠️  Could not rename GL workbook to '{new_name}': {e}")
        print("   The original filename remains.")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted by user.")
