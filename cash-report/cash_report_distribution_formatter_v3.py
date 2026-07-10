import copy
import sys
from pathlib import Path
from openpyxl import load_workbook


def stop(msg: str):
    raise Exception(msg)


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip()


def copy_cell_format(source_cell, target_cell):
    if source_cell.has_style:
        target_cell._style = copy.copy(source_cell._style)
    target_cell.number_format = source_cell.number_format
    target_cell.font = copy.copy(source_cell.font)
    target_cell.fill = copy.copy(source_cell.fill)
    target_cell.border = copy.copy(source_cell.border)
    target_cell.alignment = copy.copy(source_cell.alignment)
    target_cell.protection = copy.copy(source_cell.protection)


def column_letter_to_index(letter):
    letter = letter.upper()
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result


def get_header_map(ws):
    """
    Returns a dictionary of:
    {
        header_text: column_index
    }
    based on row 1.
    """
    header_map = {}
    for col in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=1, column=col).value)
        if header:
            header_map[header] = col
    return header_map


def validate_required_headers(ws, required_headers):
    header_map = get_header_map(ws)
    missing = [header for header in required_headers if header not in header_map]
    if missing:
        stop(
            f"Missing required column(s) on sheet '{ws.title}': "
            + ", ".join(missing)
        )
    return header_map


def delete_sheets_except(wb, keep_sheet_names):
    existing_sheets = wb.sheetnames[:]
    missing_sheets = [sheet for sheet in keep_sheet_names if sheet not in existing_sheets]

    if missing_sheets:
        stop("Missing required sheet(s): " + ", ".join(missing_sheets))

    for sheet_name in existing_sheets:
        if sheet_name not in keep_sheet_names:
            del wb[sheet_name]


def delete_columns_simultaneously(ws, col_letters):
    """
    Deletes columns as if simultaneously by deleting from right to left.
    """
    col_indexes = sorted(
        [column_letter_to_index(col) for col in col_letters],
        reverse=True
    )
    for col_idx in col_indexes:
        ws.delete_cols(col_idx, 1)


def reorder_columns_by_header(ws, ordered_headers):
    """
    Reorders worksheet columns based on row 1 headers.
    Only the requested ordered headers are retained in the new left-side layout.
    """
    header_map = validate_required_headers(ws, ordered_headers)

    max_row = ws.max_row
    max_col = ws.max_column

    # Save values
    original_values = []
    for r in range(1, max_row + 1):
        row_data = []
        for c in range(1, max_col + 1):
            row_data.append(ws.cell(r, c).value)
        original_values.append(row_data)

    # Build reordered values
    reordered_values = []
    for r in range(max_row):
        new_row = []
        for header in ordered_headers:
            source_col = header_map[header]
            new_row.append(original_values[r][source_col - 1])
        reordered_values.append(new_row)

    # Clear existing values
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).value = None

    # Write reordered values
    for r_idx, row_data in enumerate(reordered_values, start=1):
        for c_idx, value in enumerate(row_data, start=1):
            ws.cell(r_idx, c_idx).value = value


def delete_extra_columns_after(ws, keep_until_col_index):
    if ws.max_column > keep_until_col_index:
        ws.delete_cols(keep_until_col_index + 1, ws.max_column - keep_until_col_index)


def format_date_columns_as_short_date(ws):
    """
    Formats any column with header 'Date' as Excel short date
    for all data rows beneath the header.
    """
    header_map = get_header_map(ws)
    date_col = header_map.get("Date")
    if not date_col:
        return

    for r in range(2, ws.max_row + 1):
        ws.cell(r, date_col).number_format = "mm/dd/yyyy"


def format_all_rows_like_row_2(ws):
    """
    Formats every row from row 2 through max row
    using the formatting from row 2.
    """
    if ws.max_row < 2:
        return

    max_col = ws.max_column

    for r in range(2, ws.max_row + 1):
        for c in range(1, max_col + 1):
            source = ws.cell(2, c)
            target = ws.cell(r, c)
            copy_cell_format(source, target)


def format_rows_like_row_2_until_internal_id(ws):
    """
    Formats rows starting at row 2 until a later row in column A
    contains 'Internal ID'. That row is not reformatted.
    """
    if ws.max_row < 2:
        return

    max_col = ws.max_column
    stop_row = ws.max_row + 1

    for r in range(3, ws.max_row + 1):
        col_a_value = normalize_header(ws.cell(r, 1).value)
        if col_a_value == "Internal ID":
            stop_row = r
            break

    for r in range(2, stop_row):
        for c in range(1, max_col + 1):
            source = ws.cell(2, c)
            target = ws.cell(r, c)
            copy_cell_format(source, target)


def process_sgna_workbook(wb):
    keep_sheets = [
        "PP_SG&A_Pivots",
        "PP_Detail_PP_SGA",
    ]
    delete_sheets_except(wb, keep_sheets)

    ws = wb["PP_Detail_PP_SGA"]

    # Two new columns were inserted after J:
    #   K = Cash Impact
    #   L = Excluded
    # Any original delete target after J shifts +2.
    # Original: B, C, I, J, K, L, Q, R, S, T, V
    # New:      B, C, I, J, K, L, M, N, S, T, U, V, X
    delete_columns_simultaneously(
        ws,
        ["B", "C", "I", "J", "K", "L", "M", "N", "S", "T", "U", "V", "X"]
    )

    ordered_headers = [
        "Internal ID",
        "Department",
        "GL #",
        "Account",
        "Name",
        "Amount",
        "Date",
        "Month",
        "Day",
        "Year",
        "Transaction Number",
        "Document Number",
        "Check/Deposit",
        "Memo+B:O",
        "Description",
        "GL Sub-Type",
    ]

    reorder_columns_by_header(ws, ordered_headers)

    # 16 total columns kept through column P
    delete_extra_columns_after(ws, 16)

    format_all_rows_like_row_2(ws)
    format_date_columns_as_short_date(ws)


def process_cip_dev_workbook(wb):
    keep_sheets = [
        "PP_Detail_PP_Dev_CIP_Pivot",
        "PP_Detail_PP_Dev_&_CIP",
        "PP_Dev_Pivots",
        "PP_Detail_PP_Dev",
    ]
    delete_sheets_except(wb, keep_sheets)

    # Pivot sheet
    ws_pivot = wb["PP_Detail_PP_Dev_CIP_Pivot"]
    delete_columns_simultaneously(ws_pivot, ["D", "E", "F", "G"])

    detail_sheet_names = [
        "PP_Detail_PP_Dev_&_CIP",
        "PP_Detail_PP_Dev",
    ]

    ordered_headers = [
        "Internal ID",
        "PP Projects",
        "External ID",
        "GL #",
        "Account",
        "Name",
        "Amount",
        "Date",
        "Month",
        "Day",
        "Year",
        "Transaction Number",
        "Document Number",
        "Check/Deposit",
        "GL Sub-Type",
    ]

    for sheet_name in detail_sheet_names:
        ws = wb[sheet_name]

        # Two new columns were inserted after J:
        #   K = Cash Impact
        #   L = Excluded
        # Any original delete target after J shifts +2.
        # Original: B, C, I, J, M, N, O, Q, R, S, T, V
        # New:      B, C, I, J, K, L, O, P, Q, S, T, U, V, X
        delete_columns_simultaneously(
            ws,
            ["B", "C", "I", "J", "K", "L", "O", "P", "Q", "S", "T", "U", "V", "X"]
        )

        reorder_columns_by_header(ws, ordered_headers)

        # 15 total columns kept through column O
        delete_extra_columns_after(ws, 15)

        format_rows_like_row_2_until_internal_id(ws)
        format_date_columns_as_short_date(ws)


def main():
    script_folder = Path(__file__).resolve().parent

    file_name = input(
        "What is the excel file you would like to reformat for distribution?\n"
    ).strip()

    if not file_name:
        stop("No file name was entered.")

    if not file_name.lower().endswith(".xlsx"):
        file_name += ".xlsx"

    input_path = script_folder / file_name

    if not input_path.exists():
        stop(f"Input file not found in script folder: {input_path}")

    original_file_name = input_path.name
    sgna_output_path = script_folder / f"SGNA_{original_file_name}"
    cip_dev_output_path = script_folder / f"CIP_Dev_{original_file_name}"

    # Create and process SGNA copy
    wb_sgna = load_workbook(input_path)
    process_sgna_workbook(wb_sgna)
    wb_sgna.save(sgna_output_path)

    # Create and process CIP_Dev copy
    wb_cip = load_workbook(input_path)
    process_cip_dev_workbook(wb_cip)
    wb_cip.save(cip_dev_output_path)

    print("\nSuccess.")
    print(f"Created: {sgna_output_path.name}")
    print(f"Created: {cip_dev_output_path.name}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\nERROR: {e}")
        sys.exit(1)
