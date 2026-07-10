#!/usr/bin/env python3
"""
Cash Report Master Script (v23 - v22 plus source sheet name in Merge_PP_ESR Column V)
- Merge Data_Sheets -> Merge_PP_ESR (values-only, A..Y; clears A2:Y down first; overwrites V with source sheet name)
- Distribute to Reporting_Sheets via staged lookup using '<' as the separator for ALL list fields:
    B: include on shifted source Column R (original Column P; exact tokens, '<'-separated)
    C: include on Column G (exact tokens, '<'-separated)
    D: include on FIRST 3 of Column G (tokens, '<'-separated)
    E: DISQUALIFY on FIRST 3 of G (tokens, '<'-separated)
    F: SPECIAL section triggers (tokens, '<'-separated); match ANY token against G or shifted source R
- Respect the normal-section boundary above the row whose Column A == "Internal ID"
- Special section:
    * Clear old values below "Internal ID" (keep header & formatting)
    * Ensure capacity, copy formatting, write qualifying rows
- De-duplication:
    * Any row added to special is removed from that sheet’s normal section
- Date prompt (MM/DD/YYYY) and stamp Z..AC:
    Z: date (mm/dd/yyyy), AA: MM (text), AB: DD (text), AC: YYYY (text)
- Update Checkers!C129:C131 to SUM the special_handling_section (below "Internal ID")
  of each sheet named in B129:B131, summing Column H by default.
- FINAL STEP: Refresh pivot tables on these sheets via Excel COM (if pywin32 is available):
    * 'PP_Detail_PP_Dev_CIP_Pivot'
    * 'PP_SG&A_Pivots'
    * 'PP_Dev_Pivots'
"""

from pathlib import Path
from typing import List, Tuple, Any, Iterable
from copy import copy
from datetime import datetime, date
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

# Optional Excel COM for pivot refresh
try:
    import win32com.client as win32  # type: ignore
    HAS_WIN32 = True
except Exception:
    HAS_WIN32 = False

# ==== CONFIG ====
DATA_SHEETS = [
    "PP_Detail_Report",
    "PPM_Detail_Report",
    "PPS_CPB_Detail_Report",
    "PPS_Detail_Report",
    "ESR_Detail_Report",
]

SOURCE_SHEET = "Merge_PP_ESR"

MERGE_START_COL = 1   # A
MERGE_END_COL   = 25  # Y (original A:W plus 2 inserted columns after J)
START_ROW       = 2   # data starts on row 2
REPORT_END_COL  = 29  # AC (date stamp columns shifted right by 2)
SOURCE_NAME_COL = 22  # V - overwrite with originating detail report sheet name during merge

LOOKUP_FILE  = "Cash_Report_Lookup.xlsx"

# Bank merge list is stored in the separate lookup workbook
BANK_SHEET_TAB    = "Bank_Detail_Reports"  # tab name inside Cash_Report_Lookup.xlsx
BANK_SHEET_HEADER = "Sheet Name"           # expected at A1
LOOKUP_SHEET = "lookup"

# indices within A..Y slice (0-based into the row list we extract)
IDX_G = 6     # Column G
IDX_P = 17    # Column R (original Column P shifted right by 2 inserted columns after J)

# absolute column indexes in reporting sheets
COL_X, COL_Y, COL_Z, COL_AA = 26, 27, 28, 29  # Z, AA, AB, AC

# Pivot refresh target sheets (all will be attempted)
PIVOT_SHEETS_TO_REFRESH = [
    "PP_Detail_PP_Dev_CIP_Pivot",
    "PP_SG&A_Pivots",
    "PP_Dev_Pivots",
]

# Which column to sum in special_handling_section for Checkers formulas (H by default)
SPECIAL_SUM_COL = 8  # H
# =================


# -------- normalization helpers --------
def _num_to_key(x) -> str:
    try:
        if isinstance(x, float):
            if x.is_integer():
                return str(int(x))
            return f"{x:.15g}"
        return str(x)
    except Exception:
        return str(x)

def to_key(x) -> str:
    """Normalize: strip, remove BOM/ZWSP, uppercase, normalize numbers."""
    if x is None:
        return ""
    if isinstance(x, (int, float)):
        s = _num_to_key(x)
    else:
        s = str(x)
    s = s.replace("\ufeff", "").replace("\u200b", "").strip()
    return s.upper()

def split_tokens_by_lt(val) -> List[str]:
    """
    Split on '<' for all list-style lookup fields.
    Example: "A123 < BETA < 9001" -> ["A123", "BETA", "9001"]
    """
    if val in (None, ""):
        return []
    s = str(val)
    parts = [to_key(p) for p in s.split("<")]
    return [p for p in parts if p != ""]


# -------- generic helpers --------
def detect_single_workbook(folder: Path) -> Path:
    xlsx_files = [p for p in folder.glob("*.xlsx") if p.name != LOOKUP_FILE]
    if not xlsx_files:
        raise SystemExit("ERROR: No Excel (.xlsx) file (besides Cash_Report_Lookup.xlsx) found in this folder.")
    if len(xlsx_files) > 1:
        names = ", ".join(p.name for p in xlsx_files)
        raise SystemExit(f"ERROR: More than one Excel file found ({names}). Leave only one .xlsx beside {LOOKUP_FILE}.")
    return xlsx_files[0]

def find_last_data_row(ws: Worksheet, start_row: int, start_col: int, end_col: int) -> int:
    max_row = ws.max_row or start_row
    last = start_row - 1
    for r in range(start_row, max_row + 1):
        for c in range(start_col, end_col + 1):
            if ws.cell(row=r, column=c).value not in (None, ""):
                last = r
                break
    return last

def read_rows(ws: Worksheet, start_row: int, end_row: int, start_col: int, end_col: int):
    for row in ws.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
        values_only=True,
    ):
        yield list(row)

def clear_values_region(ws: Worksheet, start_row: int, start_col: int, end_col: int) -> None:
    max_row = ws.max_row or start_row
    for r in range(start_row, max_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).value = None



def get_bank_detail_sheet_names(lookup_path: Path) -> List[str]:
    """
    Reads Cash_Report_Lookup.xlsx -> Bank_Detail_Reports!A2:A.

    Requirements (per your rules):
      - Tab name: Bank_Detail_Reports (in the *separate* lookup workbook)
      - A1 must equal: 'Sheet Name'
      - Values: sheet names in A2 downward

    Robustness:
      - Skips blanks instead of stopping at the first one
      - Stops after MAX_BLANK_STREAK consecutive blanks (prevents scanning the entire column)
      - De-dupes using normalized keys (to_key) but preserves the original text for display/logging
    """
    if not lookup_path.exists():
        raise SystemExit(f"ERROR: Lookup file not found: {lookup_path}")

    wb_lookup = openpyxl.load_workbook(lookup_path, data_only=True)

    if BANK_SHEET_TAB not in wb_lookup.sheetnames:
        raise SystemExit(f"ERROR: Required lookup tab '{BANK_SHEET_TAB}' not found in {lookup_path.name}.")

    ws = wb_lookup[BANK_SHEET_TAB]

    header_val = ws.cell(row=1, column=1).value
    header_txt = str(header_val or "").strip()
    if header_txt != BANK_SHEET_HEADER:
        raise SystemExit(
            f"ERROR: Expected '{BANK_SHEET_TAB}'!A1 to be '{BANK_SHEET_HEADER}', got '{header_txt}'."
        )

    names: List[str] = []
    seen_keys = set()

    blank_streak = 0
    MAX_BLANK_STREAK = 20

    for r in range(2, (ws.max_row or 1) + 1):
        raw = ws.cell(row=r, column=1).value
        name = str(raw or "").strip()

        if not name:
            blank_streak += 1
            if blank_streak >= MAX_BLANK_STREAK:
                break
            continue

        blank_streak = 0

        k = to_key(name)
        if k not in seen_keys:
            names.append(name)
            seen_keys.add(k)

    if not names:
        raise SystemExit(
            f"ERROR: No sheet names found in '{BANK_SHEET_TAB}' under A1='{BANK_SHEET_HEADER}'."
        )

    return names


# -------- merge step --------

def merge_data_sheets(
    src_path: Path,
    target_sheet_name: str,
    period_date: date,
    mm_txt: str,
    dd_txt: str,
    yyyy_txt: str,
) -> openpyxl.Workbook:
    """
    Merge A..Y (values-only) from all detail sheets listed in Cash_Report_Lookup.xlsx -> Bank_Detail_Reports
    into target_sheet_name (Merge_PP_ESR).

    Also overwrites Column V on the merge sheet with the originating detail report sheet name.

    Also stamps Date/Month/Day/Year into Z..AC on the merge sheet, matching reporting sheet behavior:
      Z: date (mm/dd/yyyy), AA: MM (text), AB: DD (text), AC: YYYY (text)
    """
    wb_values = openpyxl.load_workbook(src_path, data_only=True)
    wb_write  = openpyxl.load_workbook(src_path)

    if target_sheet_name not in wb_write.sheetnames:
        raise SystemExit(f"ERROR: Target sheet '{target_sheet_name}' not found in {src_path.name}.")

    lookup_path = src_path.parent / LOOKUP_FILE
    sheet_names_to_merge = get_bank_detail_sheet_names(lookup_path)

    ws_target = wb_write[target_sheet_name]

    # Clear old values (keep formatting)
    clear_values_region(ws_target, START_ROW, MERGE_START_COL, MERGE_END_COL)  # A..Y
    clear_values_region(ws_target, START_ROW, COL_X, COL_AA)                  # Z..AC (stamps)

    write_row = START_ROW
    total_rows = 0

    # Normalize workbook sheet names for robust matching (handles hidden chars / case)
    sheet_map = {to_key(s): s for s in wb_write.sheetnames}

    for name in sheet_names_to_merge:
        key = to_key(name)
        actual = sheet_map.get(key)

        if not actual or actual not in wb_values.sheetnames:
            print(
                f"WARNING: '{name}' listed in {lookup_path.name}:{BANK_SHEET_TAB} but not found in workbook. Skipping."
            )
            continue

        ws_vals = wb_values[actual]
        last_row = find_last_data_row(wb_write[actual], START_ROW, MERGE_START_COL, MERGE_END_COL)
        if last_row < START_ROW:
            continue

        for row_vals in read_rows(ws_vals, START_ROW, last_row, MERGE_START_COL, MERGE_END_COL):
            # Write A..Y
            for offset, v in enumerate(row_vals):
                ws_target.cell(row=write_row, column=MERGE_START_COL + offset).value = v

            # Overwrite Column V with the originating detail report sheet name
            ws_target.cell(row=write_row, column=SOURCE_NAME_COL).value = actual

            # Stamp Z..AC (report-style)
            cx = ws_target.cell(row=write_row, column=COL_X)
            cx.value = period_date
            cx.number_format = "mm/dd/yyyy"
            ws_target.cell(row=write_row, column=COL_Y).value  = mm_txt
            ws_target.cell(row=write_row, column=COL_Z).value  = dd_txt
            ws_target.cell(row=write_row, column=COL_AA).value = yyyy_txt

            write_row += 1
            total_rows += 1

    print(f"Merged {total_rows} rows into {target_sheet_name}")
    return wb_write



# -------- lookup load --------
def load_lookup(lookup_path: Path, sheet_name: str):
    """
    Returns list of tuples:
      (report_sheet, trig_B, must_C_list, pref_D, disq_E, special_F)
    All list-style columns (B–F) use '<' as separator.
      trig_B: list[str]      (shifted source Column R exact matches (original P))
      must_C_list: list[str] (Column G exact matches)
      pref_D: list[str]      (FIRST 3 of G include)
      disq_E: list[str]      (FIRST 3 of G disqualify)
      special_F: list[str]   (SPECIAL triggers; match exact G or exact shifted R)
    """
    if not lookup_path.exists():
        raise SystemExit(f"ERROR: Lookup file not found: {lookup_path}")

    wb = openpyxl.load_workbook(lookup_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"ERROR: Lookup sheet '{sheet_name}' not found in {lookup_path.name}")

    ws = wb[sheet_name]
    out = []
    r = 2
    while True:
        a = ws.cell(row=r, column=1).value
        if a is None or str(a).strip() == "":
            break
        b = ws.cell(row=r, column=2).value  # R triggers (original P shifted +2) (<-separated)
        c = ws.cell(row=r, column=3).value  # exact G matches (<-separated)
        d = ws.cell(row=r, column=4).value  # FIRST 3 of G prefixes (<-separated)
        e = ws.cell(row=r, column=5).value  # DISQUALIFY FIRST 3 of G (<-separated)
        f = ws.cell(row=r, column=6).value  # SPECIAL triggers (<-separated)

        sheet_name_out = str(a).strip()
        trig_B = split_tokens_by_lt(b)
        must_C_list = split_tokens_by_lt(c)
        pref_D = split_tokens_by_lt(d)
        disq_E = split_tokens_by_lt(e)
        special_F = split_tokens_by_lt(f)

        out.append((sheet_name_out, trig_B, must_C_list, pref_D, disq_E, special_F))
        r += 1

    return out


# -------- sheet helpers --------
def find_special_header(ws: Worksheet) -> int:
    """
    Return the row index of the header 'Internal ID' in Column A (>=2).
    If not found, returns ws.max_row + 1 (i.e., no special section exists).
    """
    max_row = ws.max_row or 1
    for r in range(START_ROW, max_row + 1):
        if (ws.cell(row=r, column=1).value or "") == "Internal ID":
            return r
    return (max_row + 1)

def clear_non_special_section(ws: Worksheet) -> int:
    """
    Clear A..AC in the non-special section (rows 2..special_start-1). Keep formatting.
    Returns the special_start header row.
    """
    special_start = find_special_header(ws)
    end_clear = max(START_ROW, special_start - 1)
    if end_clear >= START_ROW:
        for r in range(START_ROW, end_clear + 1):
            for c in range(1, REPORT_END_COL + 1):
                ws.cell(row=r, column=c).value = None
    return special_start

def clear_special_section(ws: Worksheet, special_start: int) -> None:
    """
    Clear A..AC in the special section rows (special_start+1..bottom). Keep header/formatting.
    """
    max_row = ws.max_row or special_start
    start = special_start + 1
    if start <= max_row:
        for r in range(start, max_row + 1):
            for c in range(1, REPORT_END_COL + 1):
                ws.cell(row=r, column=c).value = None

def copy_row_format(ws: Worksheet, src_row: int, dest_row: int) -> None:
    """Copy styles for A..AC and row height from src_row to dest_row."""
    src_dim = ws.row_dimensions.get(src_row)
    if src_dim and src_dim.height is not None:
        ws.row_dimensions[dest_row].height = src_dim.height
    for col in range(1, REPORT_END_COL + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dest_row, column=col)
        try: dst.font = copy(src.font)
        except Exception: pass
        try: dst.fill = copy(src.fill)
        except Exception: pass
        try: dst.border = copy(src.border)
        except Exception: pass
        try: dst.alignment = copy(src.alignment)
        except Exception: pass
        dst.number_format = src.number_format
        try: dst.protection = copy(src.protection)
        except Exception: pass

def ensure_capacity_before_special(ws: Worksheet, needed_rows: int) -> None:
    """Ensure enough rows in the normal section (rows 2..special-1). Insert above special if needed."""
    special_start = find_special_header(ws)
    available = max(0, special_start - START_ROW)
    deficit = needed_rows - available
    if deficit <= 0:
        return
    ws.insert_rows(special_start, amount=deficit)
    for i in range(deficit):
        copy_row_format(ws, src_row=START_ROW, dest_row=special_start + i)

def ensure_capacity_in_special(ws: Worksheet, special_start: int, needed_rows: int) -> None:
    """Ensure enough rows in the special section (special_start+1..). Insert below header if needed."""
    max_row = ws.max_row or special_start
    current_rows = max(0, max_row - special_start)
    deficit = needed_rows - current_rows
    if deficit <= 0:
        return
    insert_at = special_start + 1
    ws.insert_rows(insert_at, amount=deficit)
    # Template row: first special row after insertion if present; else row 2.
    template_row = insert_at + deficit
    if template_row > ws.max_row:
        template_row = 2
    if (ws.cell(row=template_row, column=1).value or "") == "Internal ID":
        template_row = 2
    for i in range(deficit):
        copy_row_format(ws, src_row=template_row, dest_row=insert_at + i)


# -------- normal staged filter (B->C->D, then E) --------
def staged_filter_indices(
    all_rows: List[List[Any]],
    triggers_B: List[str],
    must_C_list: List[str],
    prefixes_D: List[str],   # include on FIRST 3 of G
    disqualify_E: List[str], # disqualify on FIRST 3 of G
) -> List[int]:
    included = set()
    order: List[int] = []

    # Stage B: Column R triggers (original P shifted +2) (exact)
    if triggers_B:
        tset = set(triggers_B)
        for i, row in enumerate(all_rows):
            if to_key(row[IDX_P]) in tset and i not in included:
                included.add(i); order.append(i)

    # Stage C: Column G exact matches (list, only if not already included)
    if must_C_list:
        cset = set(must_C_list)
        for i, row in enumerate(all_rows):
            if i in included:
                continue
            if to_key(row[IDX_G]) in cset:
                included.add(i); order.append(i)

    # Stage D: FIRST 3 of G in prefixes (only if not already included)
    if prefixes_D:
        pset = set(prefixes_D)
        for i, row in enumerate(all_rows):
            if i in included:
                continue
            g3 = to_key(row[IDX_G])[:3]
            if g3 in pset:
                included.add(i); order.append(i)

    # If B, C, D all empty -> include all rows (pre-E)
    if not triggers_B and not must_C_list and not prefixes_D:
        order = list(range(len(all_rows)))
        included = set(order)

    # Stage E: DISQUALIFY on FIRST 3 of G
    if disqualify_E:
        dq = set(disqualify_E)
        order = [
            i for i in order
            if to_key(all_rows[i][IDX_G])[:3] not in dq
        ]

    return order


# -------- distribution (normal + special) --------
def distribute_report_sheets(
    wb_write: openpyxl.Workbook,
    src_path: Path,
    period_date: date,
    mm_txt: str,
    dd_txt: str,
    yyyy_txt: str,
    source_sheet_name: str,
    lookup_file: str,
    lookup_sheet: str
) -> None:
    if source_sheet_name not in wb_write.sheetnames:
        raise SystemExit(f"ERROR: Source sheet '{source_sheet_name}' not found after merge.")

    ws_src = wb_write[source_sheet_name]
    last_row = find_last_data_row(ws_src, START_ROW, MERGE_START_COL, MERGE_END_COL)
    if last_row < START_ROW:
        print("Info: Distribute step skipped. No data found on Merge_PP_ESR (rows 2+).")
        return

    # Read A..Y from Merge_PP_ESR
    all_rows = list(read_rows(ws_src, START_ROW, last_row, MERGE_START_COL, MERGE_END_COL))

    lookup_path = src_path.parent / lookup_file
    mappings = load_lookup(lookup_path, sheet_name=lookup_sheet)

    for (sheet_name, trig_B, must_C, pref_D, disq_E, special_F) in mappings:
        if sheet_name not in wb_write.sheetnames:
            print(f"[skip] Reporting sheet '{sheet_name}' not found.")
            continue

        ws_target = wb_write[sheet_name]

        # ---------- Determine SPECIAL rows first (for de-dup) ----------
        special_indices: List[int] = []
        if special_F:
            trig_set = set(special_F)  # normalized tokens (from F split by '<')
            for i, row in enumerate(all_rows):
                g_val = to_key(row[IDX_G])
                p_val = to_key(row[IDX_P])
                if g_val in trig_set or p_val in trig_set:
                    special_indices.append(i)

        # ---------- NORMAL SECTION ----------
        normal_indices = staged_filter_indices(all_rows, trig_B, must_C, pref_D, disq_E)
        if special_indices:
            sset = set(special_indices)
            normal_indices = [i for i in normal_indices if i not in sset]

        # Clear normal section and ensure capacity
        special_header_row = clear_non_special_section(ws_target)
        ensure_capacity_before_special(ws_target, needed_rows=len(normal_indices))
        # recompute in case inserts shifted the header
        special_header_row = find_special_header(ws_target)
        max_allowed_row = special_header_row - 1

        # Write NORMAL rows (A..Y) + date stamps Z..AC
        write_row = START_ROW
        for idx in normal_indices:
            if write_row > max_allowed_row:
                break
            row_vals = all_rows[idx]
            for offset, v in enumerate(row_vals):
                ws_target.cell(row=write_row, column=MERGE_START_COL + offset).value = v
            cx = ws_target.cell(row=write_row, column=COL_X);  cx.value = period_date; cx.number_format = "mm/dd/yyyy"
            cy = ws_target.cell(row=write_row, column=COL_Y);  cy.value = mm_txt;     cy.number_format = "@"
            cz = ws_target.cell(row=write_row, column=COL_Z);  cz.value = dd_txt;     cz.number_format = "@"
            ca = ws_target.cell(row=write_row, column=COL_AA); ca.value = yyyy_txt;   ca.number_format = "@"
            write_row += 1

        # ---------- SPECIAL SECTION ----------
        if special_indices:
            hdr = find_special_header(ws_target)
            if hdr > ws_target.max_row:
                print(f"[info] '{sheet_name}': no 'Internal ID' header; special section skipped.")
                continue

            # Clear old special rows, ensure capacity, then write
            clear_special_section(ws_target, hdr)
            ensure_capacity_in_special(ws_target, hdr, needed_rows=len(special_indices))

            write_row = hdr + 1
            for idx in special_indices:
                row_vals = all_rows[idx]
                for offset, v in enumerate(row_vals):
                    ws_target.cell(row=write_row, column=MERGE_START_COL + offset).value = v
                cx = ws_target.cell(row=write_row, column=COL_X);  cx.value = period_date; cx.number_format = "mm/dd/yyyy"
                cy = ws_target.cell(row=write_row, column=COL_Y);  cy.value = mm_txt;     cy.number_format = "@"
                cz = ws_target.cell(row=write_row, column=COL_Z);  cz.value = dd_txt;     cz.number_format = "@"
                ca = ws_target.cell(row=write_row, column=COL_AA); ca.value = yyyy_txt;   ca.number_format = "@"
                write_row += 1

        # Uncomment for quick console feedback:
        # print(f"[done] {sheet_name}: normal={len(normal_indices)}, special={len(special_indices)}")


# -------- Update Checkers formulas (C129:C131) --------
def update_checkers_formulas_for_special_sums(wb: openpyxl.Workbook, sum_col_index: int = SPECIAL_SUM_COL) -> None:
    """
    For rows 129..131 on the 'Checkers' sheet:
      - Read the target reporting sheet name from column B
      - Find the 'Internal ID' header row on that sheet
      - Set C{row} to SUM of the special section (rows header+1 .. bottom) for the given column
        e.g., =SUM('Report'!H{header+1}:H1048576) when sum_col_index=8
    If the sheet is missing or no 'Internal ID' header is found, set 0.
    """
    if "Checkers" not in wb.sheetnames:
        print("[Checkers] Sheet 'Checkers' not found; skipping formula update.")
        return

    ws_chk = wb["Checkers"]
    col_letter = get_column_letter(sum_col_index)

    for r in range(129, 132):  # 129, 130, 131
        target_name = ws_chk.cell(row=r, column=2).value  # column B
        cell_out = ws_chk.cell(row=r, column=3)           # column C

        if not target_name or str(target_name).strip() == "":
            cell_out.value = 0
            continue

        target_name = str(target_name).strip()
        if target_name not in wb.sheetnames:
            cell_out.value = 0
            continue

        ws_tgt = wb[target_name]
        # Find the special header row ("Internal ID" in column A)
        hdr_row = find_special_header(ws_tgt)
        if hdr_row > ws_tgt.max_row:
            # no header found
            cell_out.value = 0
            continue

        start_row = hdr_row + 1
        formula = f"=SUM('{target_name}'!{col_letter}{start_row}:{col_letter}1048576)"
        cell_out.value = formula


# -------- Excel COM Pivot Refresh (multi-sheet) --------
def refresh_pivots_with_excel_multi(path: Path, sheet_names: List[str]) -> None:
    """
    Use Excel COM (pywin32) to open the workbook, refresh pivot tables
    on a list of sheets, and save.
    - If a sheet isn't found or has no pivots, falls back to RefreshAll for the workbook.
    - If pywin32 or Excel isn't available, prints a note and returns gracefully.
    """
    if not HAS_WIN32:
        print("Pivot refresh skipped (pywin32 not available). To enable: pip install pywin32")
        return

    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(str(path))

        for sheet_name in sheet_names:
            try:
                ws = wb.Worksheets(sheet_name)
            except Exception:
                print(f"[pivot] Sheet '{sheet_name}' not found; performing RefreshAll.")
                wb.RefreshAll()
                try:
                    excel.CalculateUntilAsyncQueriesDone()
                except Exception:
                    pass
                continue

            # Try to refresh pivots on this sheet
            try:
                pt_count = ws.PivotTables().Count
            except Exception:
                pt_count = 0

            if pt_count > 0:
                for i in range(1, pt_count + 1):
                    try:
                        ws.PivotTables(i).RefreshTable()
                    except Exception:
                        pass
                print(f"[pivot] Refreshed {pt_count} pivot(s) on '{sheet_name}'.")
            else:
                print(f"[pivot] No pivots detected on '{sheet_name}'; performing RefreshAll fallback.")
                wb.RefreshAll()
                try:
                    excel.CalculateUntilAsyncQueriesDone()
                except Exception:
                    pass

        wb.Save()
        wb.Close(SaveChanges=True)
        excel.Quit()
        print("Pivot refresh complete.")
    except Exception as e:
        print(f"Pivot refresh skipped due to COM error: {e}")


# -------- prompt + orchestration --------
def prompt_period_start_date() -> Tuple[date, str, str, str]:
    while True:
        raw = input("Enter the first day of the reporting period (MM/DD/YYYY): ").strip()
        try:
            dt = datetime.strptime(raw, "%m/%d/%Y").date()
            mm = f"{dt.month:02d}"
            dd = f"{dt.day:02d}"
            yyyy = f"{dt.year:04d}"
            return dt, mm, dd, yyyy
        except ValueError:
            print("Invalid date. Please use MM/DD/YYYY (e.g., 07/01/2025).")

def main():
    folder = Path(__file__).parent
    src_path = detect_single_workbook(folder)

    # date prompt
    period_date, mm_txt, dd_txt, yyyy_txt = prompt_period_start_date()

    # merge A..Y into Merge_PP_ESR
    wb_write = merge_data_sheets(src_path, SOURCE_SHEET, period_date, mm_txt, dd_txt, yyyy_txt)

    # distribute (normal + special) and stamp Z..AC
    distribute_report_sheets(
        wb_write,
        src_path,
        period_date,
        mm_txt,
        dd_txt,
        yyyy_txt,
        SOURCE_SHEET,
        LOOKUP_FILE,
        LOOKUP_SHEET,
    )

    # update 'Checkers' sheet formulas for rows 129..131 (sum special section col H by default)
    update_checkers_formulas_for_special_sums(wb_write, sum_col_index=SPECIAL_SUM_COL)

    out_path = src_path.with_name(f"{src_path.stem}_reported{src_path.suffix}")
    wb_write.save(out_path)
    print(f"Saved {out_path}")

    # Final step: refresh the pivot tables on the requested sheets
    refresh_pivots_with_excel_multi(out_path, PIVOT_SHEETS_TO_REFRESH)


if __name__ == "__main__":
    main()
