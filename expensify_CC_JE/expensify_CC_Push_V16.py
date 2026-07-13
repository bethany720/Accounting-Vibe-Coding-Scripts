from pathlib import Path
from shutil import copy2
import re
import sys
from decimal import Decimal, InvalidOperation
from datetime import datetime
from openpyxl import load_workbook

SHEET_NAME = "Transactions"
PPS_PUSH_SHEET_NAME = "PPS Push to ESR"
ESR_PUSH_SHEET_NAME = "ESR Push for Exp"
OH_PUSH_SHEET_NAME = "OH Push for Exp"

VALID_EXTENSIONS = {".xlsx"}
SCRIPT_FILENAME = Path(__file__).name
OUTPUT_PREFIX = "PPS_OH_ESR_Expensify_CC_PUSH_"

PUSH_HEADERS = [
    "Entry No.",
    "subsidiary",
    "memo",
    "trandate",
    "journalItemLine_account",
    "journalItemLine_debitAmount",
    "journalItemLine_creditAmount",
    "journalItemLine_memo",
    "constituent_project",
    "constituent_department",
]


def prompt_non_empty(prompt_text: str) -> str:
    while True:
        value = input(prompt_text).strip()
        if value:
            return value
        print("Input cannot be blank.")


def prompt_numeric_id(prompt_text: str) -> int:
    while True:
        entered_value = input(prompt_text).strip()
        if entered_value.isdigit():
            return int(entered_value)
        print("Value must be numeric with no decimals.")


def sanitize_filename_part(value: str) -> str:
    cleaned = re.sub(r'[\/:*?"<>|]+', "_", value.strip())
    cleaned = re.sub(r"\s+", "_", cleaned)
    return cleaned


def find_only_excel_file(script_dir: Path) -> Path:
    excel_files = [
        path
        for path in script_dir.iterdir()
        if path.is_file()
        and path.suffix.lower() in VALID_EXTENSIONS
        and path.name != SCRIPT_FILENAME
        and not path.name.startswith("~$")
    ]

    if len(excel_files) == 0:
        raise FileNotFoundError("No .xlsx file was found in the same folder as the Python script.")
    if len(excel_files) > 1:
        raise FileExistsError(
            "More than one .xlsx file was found in the same folder as the Python script. "
            "Please keep only one .xlsx file in that folder."
        )

    return excel_files[0]


def get_output_file(script_dir: Path, je_month: str) -> Path:
    output_name = f"{OUTPUT_PREFIX}{sanitize_filename_part(je_month)}.xlsx"
    return script_dir / output_name


def to_decimal(value) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def excel_date_text(value) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%m/%d/%Y")
    text = str(value).strip()
    if text:
        try:
            parsed = datetime.fromisoformat(text)
            return parsed.strftime("%m/%d/%Y")
        except ValueError:
            pass
    return text


def transaction_memo_value(ws, row: int) -> str:
    c_val = ws[f"C{row}"].value
    f_val = excel_date_text(ws[f"F{row}"].value)
    g_val = ws[f"G{row}"].value
    k_val = ws[f"K{row}"].value
    h_val = ws[f"H{row}"].value
    parts = [
        "" if c_val in (None, "") else str(c_val),
        f_val,
        "" if g_val in (None, "") else str(g_val),
        "" if k_val in (None, "") else str(k_val),
        "" if h_val in (None, "") else str(h_val),
    ]
    return " ".join(parts)


def get_transaction_rows(transactions_ws):
    transaction_rows = []
    row = 2
    total_l = Decimal("0")
    total_m = Decimal("0")

    while True:
        if transactions_ws[f"A{row}"].value in (None, ""):
            break

        col_l = to_decimal(transactions_ws[f"L{row}"].value)
        col_m = to_decimal(transactions_ws[f"M{row}"].value)
        total_l += col_l
        total_m += col_m

        transaction_rows.append(
            {
                "line_account": transactions_ws[f"D{row}"].value,
                "debit_from_l": transactions_ws[f"L{row}"].value if col_l != 0 else "",
                "credit_from_m": transactions_ws[f"M{row}"].value if col_m != 0 else "",
                "debit_from_m": transactions_ws[f"M{row}"].value if col_m != 0 else "",
                "credit_from_l": transactions_ws[f"L{row}"].value if col_l != 0 else "",
                "line_memo": transaction_memo_value(transactions_ws, row),
                "constituent_project": transactions_ws[f"J{row}"].value,
                "constituent_department": transactions_ws[f"R{row}"].value,
            }
        )
        row += 1

    return transaction_rows, total_l, total_m


def write_headers(ws):
    for col_idx, header in enumerate(PUSH_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)


def format_amount_cell(cell):
    if cell.value not in (None, ""):
        cell.number_format = "#,##0.00"


def build_pps_push_sheet(wb, transactions_ws, je_date: str, je_month: str) -> None:
    if PPS_PUSH_SHEET_NAME in wb.sheetnames:
        del wb[PPS_PUSH_SHEET_NAME]

    push_ws = wb.create_sheet(PPS_PUSH_SHEET_NAME)
    write_headers(push_ws)

    entry_no = f"JE{je_date} PPS Push to ESR"
    memo = f"{je_date} Push PPS Project Expenses to ESR For {je_month} Activity"
    transaction_rows, total_l, total_m = get_transaction_rows(transactions_ws)

    output_row = 2
    for item in transaction_rows:
        push_ws.cell(row=output_row, column=1, value=entry_no)
        push_ws.cell(row=output_row, column=2, value="69")
        push_ws.cell(row=output_row, column=3, value=memo)
        push_ws.cell(row=output_row, column=4, value=je_date)
        push_ws.cell(row=output_row, column=5, value=item["line_account"])

        debit_cell = push_ws.cell(row=output_row, column=6, value=item["debit_from_m"])
        credit_cell = push_ws.cell(row=output_row, column=7, value=item["credit_from_l"])
        format_amount_cell(debit_cell)
        format_amount_cell(credit_cell)

        push_ws.cell(row=output_row, column=8, value=f"To record Push from PPS to ESR for {je_month} Activity")
        push_ws.cell(row=output_row, column=9, value=item["constituent_project"])
        push_ws.cell(row=output_row, column=10, value=item["constituent_department"])
        output_row += 1

    balancing_amount = total_l - total_m
    push_ws.cell(row=output_row, column=1, value=entry_no)
    push_ws.cell(row=output_row, column=2, value="69")
    push_ws.cell(row=output_row, column=3, value=memo)
    push_ws.cell(row=output_row, column=4, value=je_date)
    push_ws.cell(row=output_row, column=5, value=21000)

    bal_debit = push_ws.cell(row=output_row, column=6, value=float(balancing_amount))
    format_amount_cell(bal_debit)

    push_ws.cell(row=output_row, column=7, value="")
    push_ws.cell(row=output_row, column=8, value=f"To record Push from PPS to ESR for {je_month} Activity")
    push_ws.cell(row=output_row, column=9, value="")
    push_ws.cell(row=output_row, column=10, value=202)


def build_esr_push_sheet(wb, transactions_ws, je_date: str, je_month: str) -> None:
    if ESR_PUSH_SHEET_NAME in wb.sheetnames:
        del wb[ESR_PUSH_SHEET_NAME]

    push_ws = wb.create_sheet(ESR_PUSH_SHEET_NAME)
    write_headers(push_ws)

    entry_no = f"JE{je_date} ESR Push for Expensify activity"
    memo = f"{je_date} Push PPS Project Expenses to ESR for {je_month} Activity"
    transaction_rows, total_l, total_m = get_transaction_rows(transactions_ws)

    output_row = 2
    for item in transaction_rows:
        push_ws.cell(row=output_row, column=1, value=entry_no)
        push_ws.cell(row=output_row, column=2, value="11")
        push_ws.cell(row=output_row, column=3, value=memo)
        push_ws.cell(row=output_row, column=4, value=je_date)
        push_ws.cell(row=output_row, column=5, value=item["line_account"])

        debit_cell = push_ws.cell(row=output_row, column=6, value=item["debit_from_l"])
        credit_cell = push_ws.cell(row=output_row, column=7, value=item["credit_from_m"])
        format_amount_cell(debit_cell)
        format_amount_cell(credit_cell)

        push_ws.cell(row=output_row, column=8, value=f"To record Push from PPS to ESR for {je_month} Activity")
        push_ws.cell(row=output_row, column=9, value=item["constituent_project"])
        push_ws.cell(row=output_row, column=10, value=item["constituent_department"])
        output_row += 1

    balancing_amount = total_l - total_m
    push_ws.cell(row=output_row, column=1, value=entry_no)
    push_ws.cell(row=output_row, column=2, value="11")
    push_ws.cell(row=output_row, column=3, value=memo)
    push_ws.cell(row=output_row, column=4, value=je_date)
    push_ws.cell(row=output_row, column=5, value=31100)

    bal_credit = push_ws.cell(row=output_row, column=7, value=float(balancing_amount))
    format_amount_cell(bal_credit)

    push_ws.cell(row=output_row, column=6, value="")
    push_ws.cell(row=output_row, column=8, value=f"To record Push from PPS to ESR for {je_month} Activity")
    push_ws.cell(row=output_row, column=9, value="")
    push_ws.cell(row=output_row, column=10, value=202)


def build_oh_push_sheet(wb, transactions_ws, je_date: str, je_month: str) -> None:
    if OH_PUSH_SHEET_NAME in wb.sheetnames:
        del wb[OH_PUSH_SHEET_NAME]

    push_ws = wb.create_sheet(OH_PUSH_SHEET_NAME)
    write_headers(push_ws)

    _, total_l, total_m = get_transaction_rows(transactions_ws)
    delta = total_l - total_m

    entry_no = f"JE{je_date}PPS push to Op Husky"
    memo = f"Pushdown of ESR's Project Expensify costs from PPS to ESR via OH for {je_month}"

    # Row 2
    push_ws.cell(row=2, column=1, value=entry_no)
    push_ws.cell(row=2, column=2, value="10")
    push_ws.cell(row=2, column=3, value=memo)
    push_ws.cell(row=2, column=4, value=je_date)
    push_ws.cell(row=2, column=5, value=21470)

    credit2 = push_ws.cell(row=2, column=7, value=float(delta))
    format_amount_cell(credit2)

    push_ws.cell(row=2, column=6, value="")
    push_ws.cell(row=2, column=8, value=memo)
    push_ws.cell(row=2, column=9, value="")
    push_ws.cell(row=2, column=10, value=202)

    # Row 3
    push_ws.cell(row=3, column=1, value=entry_no)
    push_ws.cell(row=3, column=2, value="10")
    push_ws.cell(row=3, column=3, value=memo)
    push_ws.cell(row=3, column=4, value=je_date)
    push_ws.cell(row=3, column=5, value=16020)

    debit3 = push_ws.cell(row=3, column=6, value=float(delta))
    format_amount_cell(debit3)

    push_ws.cell(row=3, column=7, value="")
    push_ws.cell(row=3, column=8, value=memo)
    push_ws.cell(row=3, column=9, value="")
    push_ws.cell(row=3, column=10, value=202)


def print_row_details(ws, row: int) -> None:
    print("This row includes the below information:")
    print(f"B: {ws[f'B{row}'].value}")
    print(f"C: {ws[f'C{row}'].value}")
    print(f"F: {ws[f'F{row}'].value}")
    print(f"G: {ws[f'G{row}'].value}")
    print(f"H: {ws[f'H{row}'].value}")
    print(f"K: {ws[f'K{row}'].value}")
    print(f"L: {ws[f'L{row}'].value}")
    print(f"M: {ws[f'M{row}'].value}")


def main() -> None:
    je_date = prompt_non_empty("What is the JE Date (MM/DD/YYYY): ")
    je_month = prompt_non_empty("What is the Reporting Period (Ex: January 2026) ")

    script_dir = Path(__file__).resolve().parent
    source_file = find_only_excel_file(script_dir)
    output_file = get_output_file(script_dir, je_month)

    copy2(source_file, output_file)

    wb = load_workbook(output_file)
    ws = wb.worksheets[0]
    ws.title = SHEET_NAME

    row = 2
    while True:
        cell_a = ws[f"A{row}"].value
        if cell_a in (None, ""):
            break

        col_d_value = ws[f"D{row}"].value
        col_d_text = "" if col_d_value is None else str(col_d_value).strip()

        if not col_d_text.startswith("12"):
            fixed_asset_gl = ws[f"B{row}"].value
            ws.delete_rows(row, 1)
            print(f"deleted row that had Fixed Asset GL: {fixed_asset_gl}")
            continue

        row += 1

    transactions_end = ws.max_row
    for row in range(2, transactions_end + 1):
        if ws[f"A{row}"].value in (None, ""):
            break

        dept_cell = ws[f"R{row}"]
        if dept_cell.value in (None, ""):
            print(f"There is not a department in row {row}")
            print_row_details(ws, row)
            dept_cell.value = prompt_numeric_id(
                "Please provide the correct department internal ID for this expense:"
            )

    for row in range(2, transactions_end + 1):
        if ws[f"A{row}"].value in (None, ""):
            break

        project_cell = ws[f"J{row}"]
        if project_cell.value in (None, ""):
            print(f"There is not a PROJECT in row {row}")
            print_row_details(ws, row)
            project_cell.value = prompt_numeric_id(
                "Please provide the correct project internal ID for this expense:"
            )

    ws.insert_cols(19, 1)
    row = 2
    while True:
        if ws[f"A{row}"].value in (None, ""):
            break
        ws[f"S{row}"] = (
            f'=CONCATENATE(C{row}," ",TEXT(F{row},"MM/DD/YYYY")," ",G{row}," ",K{row}," ",H{row})'
        )
        row += 1

    build_pps_push_sheet(wb, ws, je_date, je_month)
    build_esr_push_sheet(wb, ws, je_date, je_month)
    build_oh_push_sheet(wb, ws, je_date, je_month)

    from csv import writer as csv_writer

    def sheet_to_csv(ws_obj, csv_path):
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv_writer(f)
            for row in ws_obj.iter_rows(values_only=True):
                w.writerow(["" if v is None else v for v in row])

    base_dir = Path(output_file).parent
    for sheet_name in [PPS_PUSH_SHEET_NAME, ESR_PUSH_SHEET_NAME, OH_PUSH_SHEET_NAME]:
        if sheet_name in wb.sheetnames:
            csv_name = f"{sheet_name}_{sanitize_filename_part(je_month)}.csv"
            sheet_to_csv(wb[sheet_name], base_dir / csv_name)

    wb.save(output_file)

    print(f"JE_Date entered: {je_date}")
    print(f"JE_Month entered: {je_month}")
    print(f"Source file used: {source_file.name}")
    print(f"Saved updated workbook as: {output_file.name}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"Error: {exc}")
        sys.exit(1)
